from django.shortcuts import render, get_object_or_404, redirect
from .models import (
    AccountLedger,
    Car,
    CarEvaluation,
    CarHistory,
    CarLocation,
    CarReservation,
    Sale,
    SaleInstallment,
    Notification,
    Customer,
    CustomerAccount,
    Currency,
    Expense,
    GeneralExpense,
    CarDocument,
    DebtPayment,
    Employee,
    EmployeeCommission,
    EmployeeRole,
    ExchangeRate,
    FinanceVoucher,
    InventoryTransaction,
    Invoice,
    InvoiceLine,
    OperationLog,
    AuditLog,
    InterfaceAccess,
    PlatformTenant,
    SalesCommission,
    CarMaintenance,
    GlobalAuditLog,
    DailyClosing,
    FinancialAccount,
    FinancialContainer,
    JournalEntry,
    JournalEntryLine,
    FiscalPeriodClosing,
    FiscalUnlockRequest,
    Supplier,
    SupplierInvoice,
    SupplierPayment,
    TenantUserGoogleIdentity,
    TaxRate,
    UserThemePreference,
)
from django.db.models import Sum, F, Q
from django.db import transaction, IntegrityError, connections
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, FileResponse, JsonResponse
from django.contrib.sessions.models import Session
import csv
import hashlib
import json
import openpyxl
import secrets
import urllib.error
import urllib.parse
import urllib.request
from .forms import (
    SaleForm,
    CarForm,
    DebtPaymentForm,
    CarMaintenanceForm,
    CarDocumentForm,
    GeneralExpenseForm,
    ReceiptVoucherForm,
    PaymentVoucherForm,
    OperatingExpenseVoucherForm,
    TenantLoginForm,
    TenantRegisterForm,
    PlatformOwnerLoginForm,
    TenantSwitchForm,
    DailyClosingForm,
    BankReconciliationUploadForm,
    FinancialAccountForm,
    FinancialContainerForm,
    FiscalPeriodClosingForm,
    FiscalUnlockRequestForm,
)
from .consistency_checks import build_financial_consistency_report
from .accounting import (
    build_trial_balance_rows,
    ensure_default_chart_of_accounts,
    ensure_default_financial_containers,
    get_default_financial_container,
    sync_journal_entry_for_voucher,
)
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.files.storage import default_storage
from django.core.cache import cache
from django.conf import settings
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from datetime import datetime, timedelta
import re
from decimal import Decimal
from pathlib import Path
from collections import defaultdict
from functools import wraps
from django.contrib.auth import logout, login
from django.contrib import messages
from django.contrib.auth.models import User
from django.urls import reverse
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_http_methods, require_POST
from .tenant_database import migrate_tenant_database, ensure_tenant_connection, normalize_tenant_id
from .tenant_context import clear_current_tenant, set_current_tenant, get_current_tenant_db_alias
from .tenant_registry import get_cached_tenant_metadata, is_valid_tenant_access_key
from .platform_audit import write_platform_audit
from .services import ReportService, SalesService

PERMISSION_FIELDS = [
    'can_access_dashboard',
    'can_access_cars',
    'can_access_reports',
    'can_access_debts',
    'can_access_timeline',
    'can_access_system_users',
    'can_add_maintenance_expenses',
]

PLATFORM_OWNER_SESSION_KEY = 'platform_owner_authenticated'
PLATFORM_OWNER_USERNAME_KEY = 'platform_owner_username'
TENANT_DB_ALIAS_SESSION_KEY = 'tenant_db_alias'
GOOGLE_OAUTH_STATE_SESSION_KEY = 'google_oauth_state'
PENDING_TENANT_LOGIN_SESSION_KEY = 'pending_tenant_login'
PENDING_TENANT_REGISTER_SESSION_KEY = 'pending_tenant_register'
GOOGLE_FLOW_TIMEOUT_SECONDS = 600
CENTRAL_AUDIT_EVENT_LIMIT = 120
CENTRAL_AUDIT_TENANT_LOG_LIMIT = 8
HIDDEN_PLATFORM_LOGIN_WINDOW_SECONDS = int(getattr(settings, 'SECURITY_PLATFORM_LOGIN_WINDOW_SECONDS', 900))
HIDDEN_PLATFORM_LOGIN_MAX_ATTEMPTS = int(getattr(settings, 'SECURITY_PLATFORM_LOGIN_MAX_ATTEMPTS', 5))
PLATFORM_ONLY_TABLES = {
    'sales_platformtenant',
    'sales_globalauditlog',
    'sales_tenantbackuprecord',
    'sales_tenantmigrationrecord',
    'sales_userthemepreference',
}
CENTRAL_TENANT_BUSINESS_MODELS = [
    AccountLedger,
    Car,
    CarEvaluation,
    CarHistory,
    CarLocation,
    CarReservation,
    Sale,
    SaleInstallment,
    Customer,
    CustomerAccount,
    Currency,
    Expense,
    GeneralExpense,
    CarDocument,
    DebtPayment,
    Employee,
    EmployeeCommission,
    EmployeeRole,
    ExchangeRate,
    FinanceVoucher,
    InventoryTransaction,
    Invoice,
    InvoiceLine,
    OperationLog,
    AuditLog,
    InterfaceAccess,
    Notification,
    CarMaintenance,
    SalesCommission,
    FinancialAccount,
    FinancialContainer,
    JournalEntry,
    JournalEntryLine,
    FiscalPeriodClosing,
    FiscalUnlockRequest,
    Supplier,
    SupplierInvoice,
    SupplierPayment,
    TaxRate,
]


def _google_oauth_enabled():
    return bool(
        getattr(settings, 'GOOGLE_OAUTH_ENABLED', False)
        and
        getattr(settings, 'GOOGLE_OAUTH_CLIENT_ID', '').strip()
        and getattr(settings, 'GOOGLE_OAUTH_CLIENT_SECRET', '').strip()
        and getattr(settings, 'GOOGLE_OAUTH_REDIRECT_URI', '').strip()
    )


def _build_google_authorize_url(request, flow):
    state_seed = f"{flow}:{timezone.now().timestamp()}:{secrets.token_urlsafe(18)}"
    state_token = hashlib.sha256(state_seed.encode('utf-8')).hexdigest()
    nonce_token = secrets.token_urlsafe(22)
    request.session[GOOGLE_OAUTH_STATE_SESSION_KEY] = {
        'token': state_token,
        'nonce': nonce_token,
        'flow': flow,
        'created_at': timezone.now().isoformat(),
    }
    params = {
        'client_id': settings.GOOGLE_OAUTH_CLIENT_ID,
        'response_type': 'code',
        'scope': 'openid email profile',
        'redirect_uri': settings.GOOGLE_OAUTH_REDIRECT_URI,
        'state': state_token,
        'nonce': nonce_token,
        'prompt': 'select_account',
    }
    return f"https://accounts.google.com/o/oauth2/v2/auth?{urllib.parse.urlencode(params)}"


def _is_recent_pending_payload(payload):
    if not isinstance(payload, dict):
        return False
    created_at_raw = payload.get('created_at')
    if not created_at_raw:
        return False
    try:
        created_at = datetime.fromisoformat(created_at_raw)
    except ValueError:
        return False

    if timezone.is_naive(created_at):
        created_at = timezone.make_aware(created_at, timezone.get_current_timezone())

    age_seconds = (timezone.now() - created_at).total_seconds()
    return 0 <= age_seconds <= GOOGLE_FLOW_TIMEOUT_SECONDS


def _exchange_google_code_for_tokens(code):
    data = urllib.parse.urlencode({
        'code': code,
        'client_id': settings.GOOGLE_OAUTH_CLIENT_ID,
        'client_secret': settings.GOOGLE_OAUTH_CLIENT_SECRET,
        'redirect_uri': settings.GOOGLE_OAUTH_REDIRECT_URI,
        'grant_type': 'authorization_code',
    }).encode('utf-8')
    req = urllib.request.Request(
        'https://oauth2.googleapis.com/token',
        data=data,
        headers={'Content-Type': 'application/x-www-form-urlencoded'},
        method='POST',
    )
    with urllib.request.urlopen(req, timeout=12) as response:
        return json.loads(response.read().decode('utf-8'))


def _verify_google_id_token(id_token, expected_nonce):
    params = urllib.parse.urlencode({'id_token': id_token})
    url = f"https://oauth2.googleapis.com/tokeninfo?{params}"
    with urllib.request.urlopen(url, timeout=12) as response:
        payload = json.loads(response.read().decode('utf-8'))

    if payload.get('aud') != settings.GOOGLE_OAUTH_CLIENT_ID:
        raise ValidationError('رمز Google غير مخصص لهذا التطبيق.')

    if payload.get('nonce') != expected_nonce:
        raise ValidationError('رمز Google غير صالح. أعد المحاولة.')

    if payload.get('email_verified') not in {'true', True}:
        raise ValidationError('يجب توثيق البريد الإلكتروني في حساب Google أولًا.')

    sub = (payload.get('sub') or '').strip()
    email = (payload.get('email') or '').strip().lower()
    if not sub or not email:
        raise ValidationError('تعذر قراءة بيانات حساب Google.')

    return {
        'sub': sub,
        'email': email,
        'email_verified': True,
    }


def _bind_or_validate_google_identity(*, alias, user, google_profile):
    identity = TenantUserGoogleIdentity.objects.using(alias).filter(user=user).first()
    if identity and identity.google_sub != google_profile['sub']:
        raise ValidationError('تم ربط هذا الحساب مسبقًا بحساب Google مختلف.')

    sub_owner = (
        TenantUserGoogleIdentity.objects.using(alias)
        .select_related('user')
        .filter(google_sub=google_profile['sub'])
        .first()
    )
    if sub_owner and sub_owner.user_id != user.id:
        raise ValidationError('حساب Google هذا مرتبط بمستخدم آخر داخل نفس المعرض.')

    if identity is None:
        TenantUserGoogleIdentity.objects.using(alias).create(
            user=user,
            google_sub=google_profile['sub'],
            google_email=google_profile['email'],
            email_verified=True,
        )
        return

    identity.google_email = google_profile['email']
    identity.email_verified = google_profile['email_verified']
    identity.save(using=alias, update_fields=['google_email', 'email_verified', 'last_verified_at'])


def _clear_pending_google_flow_session(request):
    request.session.pop(GOOGLE_OAUTH_STATE_SESSION_KEY, None)
    request.session.pop(PENDING_TENANT_LOGIN_SESSION_KEY, None)
    request.session.pop(PENDING_TENANT_REGISTER_SESSION_KEY, None)


def _extract_client_ip(request):
    x_forwarded_for = (request.META.get('HTTP_X_FORWARDED_FOR') or '').strip()
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()[:64]
    return (request.META.get('REMOTE_ADDR') or '').strip()[:64]


def _hidden_platform_login_cache_key(username, ip_address):
    normalized_user = (username or '').strip().lower() or '_blank_'
    normalized_ip = (ip_address or '').strip() or '_noip_'
    return f"sec:hidden-platform-login:{normalized_user}:{normalized_ip}"


def _write_hidden_platform_login_failure_audit(request, username, reason):
    ip_address = _extract_client_ip(request)
    user_agent = (request.META.get('HTTP_USER_AGENT') or '').strip()
    compact_ua = (user_agent[:90] + '...') if len(user_agent) > 93 else user_agent
    notes = f"{reason} | ip={ip_address or 'unknown'} | ua={compact_ua or 'unknown'}"
    write_platform_audit(
        event_type='platform_login_failed',
        actor_username=(username or '')[:150],
        notes=notes,
    )


def _is_platform_owner_session(request):
    return bool(request.session.get(PLATFORM_OWNER_SESSION_KEY))


def _has_platform_wide_monitor_access(request):
    if _is_platform_owner_session(request):
        return True

    user = getattr(request, 'user', None)
    if user is None or not user.is_authenticated or not user.is_superuser:
        return False

    return getattr(user._state, 'db', None) == 'default'


def _has_active_tenant_context(request):
    tenant_id = normalize_tenant_id(request.session.get('tenant_id'))
    tenant_alias = (request.session.get(TENANT_DB_ALIAS_SESSION_KEY) or '').strip()
    return bool(tenant_id and tenant_alias == f'tenant_{tenant_id}')


def _require_platform_owner_session(request):
    if not _is_platform_owner_session(request):
        raise PermissionDenied


def require_interface_access(permission_field):
    def decorator(view_func):
        @wraps(view_func)
        @login_required
        def wrapped(request, *args, **kwargs):
            if not _has_active_tenant_context(request):
                if _is_platform_owner_session(request):
                    messages.error(request, 'اختر معرضًا أولًا قبل تنفيذ عمليات المعرض.')
                    return redirect('platform_switch_tenant')
                raise PermissionDenied('لا توجد بيئة معرض مفعلة لهذا الحساب.')

            if request.user.is_superuser or _is_platform_owner_session(request):
                return view_func(request, *args, **kwargs)

            access, _ = InterfaceAccess.objects.get_or_create(user=request.user)
            if not getattr(access, permission_field, True):
                raise PermissionDenied('ليس لديك صلاحية للوصول إلى هذه العملية.')

            return view_func(request, *args, **kwargs)

        return wrapped

    return decorator


def _require_tenant_alias_or_raise(message='لا توجد بيئة معرض مفعلة لهذه العملية.'):
    tenant_alias = (get_current_tenant_db_alias() or '').strip()
    if not tenant_alias.startswith('tenant_'):
        raise PermissionDenied(message)
    return tenant_alias


def welcome(request):
    return render(request, 'sales/welcome.html')

def home(request):
    stats = {
        'available_cars': 0,
        'sold_cars': 0,
        'pending_debts': 0,
        'operations': 0,
        'active_tenants': 0,
        'system_users': 0,
        'unread_notifications': 0,
    }

    try:
        stats['available_cars'] = Car.objects.filter(is_sold=False).count()
        stats['sold_cars'] = Car.objects.filter(is_sold=True).count()
        stats['pending_debts'] = Sale.objects.filter(amount_paid__lt=F('sale_price')).count()
        stats['operations'] = OperationLog.objects.count()
        stats['active_tenants'] = PlatformTenant.objects.filter(is_active=True, is_deleted=False).count()
        stats['system_users'] = User.objects.count()
        stats['unread_notifications'] = Notification.objects.filter(is_read=False).count()
    except Exception:
        pass

    return render(request, 'sales/home.html', {'home_stats': stats})

@require_interface_access('can_access_dashboard')
def dashboard(request):
    available_cars = Car.objects.filter(is_sold=False).count()
    total_revenue = Sale.objects.aggregate(Sum('sale_price'))['sale_price__sum'] or 0
    notifications = Notification.objects.filter(is_read=False)[:5]
    maintenance_series, maintenance_current_month, maintenance_previous_avg = _build_maintenance_monthly_series(6)

    context = {
        'available_cars': available_cars,
        'total_revenue': total_revenue,
        'notifications': notifications,
        'maintenance_current_month': maintenance_current_month,
        'maintenance_previous_avg': maintenance_previous_avg,
        'maintenance_monthly_series': maintenance_series,
        'dashboard_ok_url': reverse('dashboard'),
    }
    return render(request, 'sales/dashboard.html', context)

@require_interface_access('can_access_debts')
def debts_list(request):
    # جلب جميع المبيعات التي فيها مبلغ متبقي
    # نستخدم التحقق من أن المبلغ المدفوع أقل من سعر البيع
    pending_sales = [sale for sale in Sale.objects.all() if not sale.is_fully_paid]
    
    total_debts = sum(sale.remaining_amount for sale in pending_sales)

    return render(request, 'sales/debts.html', {
        'pending_sales': pending_sales,
        'total_debts': total_debts,
        'debts_ok_url': reverse('debts_list'),
    })


@require_interface_access('can_access_debts')
def debt_aging_report(request):
    today = timezone.localdate()
    pending_sales = (
        Sale.objects.select_related('car', 'customer')
        .filter(amount_paid__lt=F('sale_price'))
        .order_by('sale_date')
    )

    bucket_specs = [
        ('not_due', 'غير مستحق بعد', None, -1),
        ('0_30', '0 - 30 يوم', 0, 30),
        ('31_60', '31 - 60 يوم', 31, 60),
        ('61_90', '61 - 90 يوم', 61, 90),
        ('91_plus', 'أكثر من 90 يوم', 91, None),
    ]

    bucket_map = {
        key: {
            'key': key,
            'label': label,
            'count': 0,
            'amount_sr': Decimal('0'),
        }
        for key, label, _min_days, _max_days in bucket_specs
    }

    rows = []
    total_outstanding_sr = Decimal('0')

    for sale in pending_sales:
        sale_day = timezone.localtime(sale.sale_date).date() if sale.sale_date else today
        due_date = sale.debt_due_date or sale_day
        age_days = (today - due_date).days
        remaining_amount = sale.remaining_amount
        remaining_amount_sr = _to_sr(remaining_amount, getattr(sale.car, 'currency', 'SR'))

        bucket_key = '91_plus'
        for key, _label, min_days, max_days in bucket_specs:
            if min_days is None and max_days is not None and age_days <= max_days:
                bucket_key = key
                break
            if max_days is None and age_days >= min_days:
                bucket_key = key
                break
            if max_days is not None and min_days <= age_days <= max_days:
                bucket_key = key
                break

        bucket_entry = bucket_map[bucket_key]
        bucket_entry['count'] += 1
        bucket_entry['amount_sr'] += remaining_amount_sr
        total_outstanding_sr += remaining_amount_sr

        rows.append({
            'sale': sale,
            'age_days': age_days,
            'due_date': due_date,
            'age_text': f"بعد {abs(age_days)} يوم" if age_days < 0 else f"{age_days} يوم",
            'bucket_label': bucket_entry['label'],
            'remaining_amount': remaining_amount,
            'remaining_amount_sr': remaining_amount_sr,
        })

    rows.sort(key=lambda item: item['age_days'], reverse=True)
    aging_buckets = [bucket_map[key] for key, _label, _min_days, _max_days in bucket_specs]

    return render(request, 'sales/debt_aging_report.html', {
        'rows': rows,
        'aging_buckets': aging_buckets,
        'total_outstanding_sr': total_outstanding_sr,
        'as_of_date': today,
    })


@require_interface_access('can_access_debts')
def add_debt_payment(request, sale_id):
    tenant_alias = _require_tenant_alias_or_raise('لا توجد بيئة معرض مفعلة لسداد المديونيات.')
    sale = get_object_or_404(
        Sale.objects.using(tenant_alias).select_related('car', 'customer'),
        id=sale_id,
    )
    ensure_default_chart_of_accounts(alias=tenant_alias)
    ensure_default_financial_containers(alias=tenant_alias)

    current_remaining = sale.remaining_amount

    last_payment = DebtPayment.objects.using(tenant_alias).order_by('-id').first()
    next_number = 1 if not last_payment else (last_payment.id + 1)
    default_receipt_number = f"SD-{next_number:05d}"
    default_payment_date = timezone.localdate()

    if request.method == 'POST':
        form = DebtPaymentForm(request.POST)
        if form.is_valid():
            paid_amount = form.cleaned_data['paid_amount']
            if paid_amount > current_remaining:
                form.add_error('paid_amount', 'المبلغ المسدد لا يمكن أن يتجاوز المبلغ المتبقي.')
            else:
                with transaction.atomic(using=tenant_alias):
                    sale = (
                        Sale.objects.using(tenant_alias)
                        .select_for_update()
                        .select_related('car', 'customer')
                        .get(pk=sale_id)
                    )

                    fresh_remaining = sale.remaining_amount
                    if paid_amount > fresh_remaining:
                        form.add_error('paid_amount', 'المبلغ المسدد لا يمكن أن يتجاوز المبلغ المتبقي.')
                        context = {
                            'sale': sale,
                            'current_remaining': fresh_remaining,
                            'form': form,
                        }
                        return render(request, 'sales/debt_payment.html', context)

                    payment = form.save(commit=False)
                    payment.sale = sale
                    if DebtPayment.objects.using(tenant_alias).filter(receipt_number=payment.receipt_number).exists():
                        form.add_error('receipt_number', 'رقم السند مستخدم مسبقاً. أدخل رقمًا مختلفًا.')
                        context = {
                            'sale': sale,
                            'current_remaining': current_remaining,
                            'form': form,
                        }
                        return render(request, 'sales/debt_payment.html', context)
                    payment.save(using=tenant_alias)

                    settlement_container = get_default_financial_container(
                        alias=tenant_alias,
                        preferred_type=FinancialContainer.TYPE_MAIN_CASH,
                        currency=sale.car.currency,
                    )
                    settlement_credit_choice = FinanceVoucher.ACCOUNT_CASH_BOX
                    if settlement_container and settlement_container.container_type == FinancialContainer.TYPE_BANK:
                        settlement_credit_choice = FinanceVoucher.ACCOUNT_BANK

                    FinanceVoucher.objects.using(tenant_alias).get_or_create(
                        voucher_number=payment.receipt_number,
                        defaults={
                            'voucher_type': 'settlement',
                            'voucher_date': payment.payment_date,
                            'person_name': sale.customer.name,
                            'amount': payment.paid_amount,
                            'currency': sale.car.currency,
                            'reason': f"تسديد مديونية سيارة {sale.car.brand} {sale.car.model_name}",
                            'linked_car': sale.car,
                            'financial_container': settlement_container,
                            'debit_account': FinanceVoucher.ACCOUNT_CASH_BOX,
                            'credit_account': settlement_credit_choice,
                        }
                    )

                    sale.amount_paid = sale.amount_paid + paid_amount
                    updated_fields = ['amount_paid']
                    if sale.amount_paid >= sale.sale_price and sale.debt_due_date is not None:
                        sale.debt_due_date = None
                        updated_fields.append('debt_due_date')
                    sale.save(using=tenant_alias, update_fields=updated_fields)

                    SalesService.allocate_payment_to_installments(
                        tenant_alias=tenant_alias,
                        sale_id=sale.pk,
                        payment_amount=paid_amount,
                    )

                return redirect(f"{reverse('debts_list')}?debt_paid=1")
    else:
        form = DebtPaymentForm(initial={
            'receipt_number': default_receipt_number,
            'payment_date': default_payment_date,
        })

    context = {
        'sale': sale,
        'current_remaining': current_remaining,
        'form': form,
    }
    return render(request, 'sales/debt_payment.html', context)

@require_interface_access('can_access_reports')
def export_sales_excel(request):
    # إنشاء كتاب عمل إكسل جديد
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تقرير المبيعات"

    # إضافة العناوين (الصف الأول)
    headers = ['السيارة', 'العميل', 'سعر البيع', 'تاريخ البيع']
    ws.append(headers)

    # جلب جميع المبيعات من قاعدة البيانات وإضافتها لملف الإكسل
    sales = Sale.objects.all()
    for sale in sales:
        ws.append([
            str(sale.car),
            str(sale.customer),
            sale.sale_price,
            sale.sale_date.strftime('%Y-%m-%d %H:%M') # تنسيق التاريخ
        ])

    # تجهيز الملف للإرسال كاستجابة للمتصفح (Download)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sales_report.xlsx"'
    wb.save(response)
    return response

@require_interface_access('can_access_cars')
def car_list(request):
    # استقبال نص البحث من المربع في الواجهة
    query = request.GET.get('search')
    
    # جلب جميع السيارات كبداية
    cars = Car.objects.all().order_by('-created_at')

    # إذا قام المستخدم بكتابة شيء في مربع البحث
    if query:
        cars = cars.filter(
            Q(brand__icontains=query) |        # البحث في الماركة
            Q(model_name__icontains=query) |   # البحث في الموديل
            Q(vin__icontains=query)            # البحث في رقم الشاسيه
        )

    return render(request, 'sales/car_list.html', {
        'cars': cars,
        'query': query,
        'car_list_ok_url': reverse('car_list'),
    })

@require_interface_access('can_access_cars')
def car_edit(request, car_id):
    car = get_object_or_404(Car, id=car_id)
    
    if request.method == 'POST':
        form = CarForm(request.POST, request.FILES, instance=car)
        if form.is_valid():
            form.save()
            return redirect(f"{reverse('car_list')}?car_updated=1")
    else:
        form = CarForm(instance=car)
        
    return render(request, 'sales/car_edit.html', {'form': form, 'car': car})

@require_interface_access('can_access_cars')
def car_detail(request, car_id):
    car = get_object_or_404(Car, id=car_id)
    maintenance_records = car.maintenance_records.select_related('added_by', 'journal_voucher').all()
    document_records = car.documents.select_related('uploaded_by').all()

    can_add_maintenance = request.user.is_superuser or _is_platform_owner_session(request)
    if not can_add_maintenance:
        access, _ = InterfaceAccess.objects.get_or_create(user=request.user)
        can_add_maintenance = access.can_add_maintenance_expenses

    selected_tab = request.GET.get('tab', 'details')
    maintenance_form = CarMaintenanceForm(initial={'operation_date': timezone.localdate()})
    document_form = CarDocumentForm()

    if request.method == 'POST':
        form_action = request.POST.get('form_action', 'maintenance')

        if form_action == 'add_document':
            selected_tab = 'documents'
            document_form = CarDocumentForm(request.POST, request.FILES)
            if document_form.is_valid():
                document = document_form.save(commit=False)
                document.car = car
                document.uploaded_by = request.user
                document.save()
                return redirect(f"{reverse('car_detail', args=[car.id])}?document_saved=1&tab=documents")
        else:
            if not can_add_maintenance:
                raise PermissionDenied('ليس لديك صلاحية لإضافة مصروفات الصيانة.')

            if car.is_sold:
                return redirect(f"{reverse('car_detail', args=[car.id])}?maintenance_error=sold&tab=maintenance")

            maintenance_form = CarMaintenanceForm(request.POST, request.FILES)
            selected_tab = 'maintenance'
            if maintenance_form.is_valid():
                maintenance = maintenance_form.save(commit=False)
                maintenance.car = car
                maintenance.added_by = request.user
                maintenance.save()
                return redirect(f"{reverse('car_detail', args=[car.id])}?maintenance_saved=1&tab=maintenance")

    today = timezone.localdate()
    insurance_days_left = None
    registration_days_left = None
    if car.insurance_expiry:
        insurance_days_left = (car.insurance_expiry - today).days
    if car.registration_expiry:
        registration_days_left = (car.registration_expiry - today).days

    return render(request, 'sales/car_detail.html', {
        'car': car,
        'maintenance_records': maintenance_records,
        'document_records': document_records,
        'maintenance_total': car.maintenance_total,
        'total_cost_price': car.total_cost_price,
        'expected_profit': car.expected_profit,
        'can_add_maintenance': can_add_maintenance,
        'maintenance_form': maintenance_form,
        'document_form': document_form,
        'insurance_days_left': insurance_days_left,
        'registration_days_left': registration_days_left,
        'selected_tab': selected_tab,
    })

@require_interface_access('can_access_cars')
def process_sale(request, car_id=None):
    tenant_alias = _require_tenant_alias_or_raise('لا توجد بيئة معرض مفعلة لإتمام عملية البيع.')
    ensure_default_chart_of_accounts(alias=tenant_alias)
    ensure_default_financial_containers(alias=tenant_alias)

    if car_id is None:
        raw_car_id = (request.POST.get('car_id') or request.GET.get('car_id') or '').strip()
        if not raw_car_id.isdigit():
            raise PermissionDenied('يلزم تحديد سيارة لإتمام عملية البيع.')
        car_id = int(raw_car_id)

    car = get_object_or_404(Car.objects.using(tenant_alias), id=car_id, is_sold=False)

    if request.method == 'POST':
        form = SaleForm(request.POST, request.FILES, tenant_alias=tenant_alias)
        if form.is_valid():
            selected_container = form.cleaned_data.get('financial_container')
            try:
                SalesService.execute_credit_sale(
                    tenant_alias=tenant_alias,
                    car_id=car.id,
                    customer_name=form.cleaned_data['customer_name'],
                    customer_phone=form.cleaned_data['customer_phone'],
                    customer_national_id=form.cleaned_data['customer_national_id'],
                    total_sale_price=form.cleaned_data['sale_price'],
                    down_payment=form.cleaned_data['amount_paid'],
                    payment_schedule=form.cleaned_data.get('payment_schedule', []),
                    debt_due_date=form.cleaned_data.get('debt_due_date'),
                    sale_contract_image=form.cleaned_data.get('sale_contract_image'),
                    actor=request.user,
                    currency_rate=form.cleaned_data.get('currency_rate'),
                    financial_container_id=selected_container.pk if selected_container else None,
                    request_path=request.path,
                    ip_address=request.META.get('REMOTE_ADDR', ''),
                    device_type='mobile' if any(token in (request.META.get('HTTP_USER_AGENT') or '').lower() for token in ['iphone', 'android', 'mobile', 'ipad']) else 'desktop',
                    browser=(
                        'Edge' if 'edg/' in (request.META.get('HTTP_USER_AGENT') or '').lower()
                        else 'Chrome' if 'chrome/' in (request.META.get('HTTP_USER_AGENT') or '').lower() and 'edg/' not in (request.META.get('HTTP_USER_AGENT') or '').lower()
                        else 'Firefox' if 'firefox/' in (request.META.get('HTTP_USER_AGENT') or '').lower()
                        else 'Safari' if 'safari/' in (request.META.get('HTTP_USER_AGENT') or '').lower() and 'chrome/' not in (request.META.get('HTTP_USER_AGENT') or '').lower()
                        else ''
                    ),
                    geo_location=(request.META.get('HTTP_CF_IPCOUNTRY') or request.META.get('HTTP_X_APPENGINE_COUNTRY') or ''),
                )
                return redirect(f"{reverse('dashboard')}?sale_created=1")
            except ValidationError as exc:
                if hasattr(exc, 'message_dict'):
                    for field_name, message_list in exc.message_dict.items():
                        for message in message_list:
                            if field_name in form.fields:
                                form.add_error(field_name, message)
                            else:
                                form.add_error(None, message)
                else:
                    for message in exc.messages:
                        form.add_error(None, message)
    else:
        form = SaleForm(tenant_alias=tenant_alias)
    
    return render(request, 'sales/process_sale.html', {'form': form, 'car': car})

@require_interface_access('can_access_reports')
def financial_reports(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    return render(request, 'sales/reports.html')


@require_interface_access('can_access_reports')
def car_profit_report(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    tenant_alias = _require_tenant_alias_or_raise('لا توجد بيئة معرض مفعلة لاستخراج تقرير أرباح السيارات.')
    rows = []

    sales = (
        Sale.objects.using(tenant_alias)
        .select_related('car', 'customer')
        .order_by('-sale_date', '-id')
    )

    total_purchase_sr = Decimal('0')
    total_cost_sr = Decimal('0')
    total_sale_sr = Decimal('0')
    total_profit_sr = Decimal('0')

    for sale in sales:
        car = sale.car
        currency = getattr(car, 'currency', 'SR')
        purchase_cost = car.cost_price or Decimal('0')
        total_cost = car.total_cost_price or Decimal('0')
        sale_price = sale.sale_price or Decimal('0')
        actual_profit = sale_price - total_cost

        purchase_cost_sr = _to_sr(purchase_cost, currency)
        total_cost_sr_value = _to_sr(total_cost, currency)
        sale_price_sr = _to_sr(sale_price, currency)
        actual_profit_sr = _to_sr(actual_profit, currency)

        total_purchase_sr += purchase_cost_sr
        total_cost_sr += total_cost_sr_value
        total_sale_sr += sale_price_sr
        total_profit_sr += actual_profit_sr

        rows.append({
            'sale': sale,
            'purchase_cost': purchase_cost,
            'total_cost': total_cost,
            'sale_price': sale_price,
            'actual_profit': actual_profit,
            'purchase_cost_sr': purchase_cost_sr,
            'total_cost_sr': total_cost_sr_value,
            'sale_price_sr': sale_price_sr,
            'actual_profit_sr': actual_profit_sr,
        })

    return render(request, 'sales/car_profit_report.html', {
        'rows': rows,
        'sold_count': len(rows),
        'total_purchase_sr': total_purchase_sr,
        'total_cost_sr': total_cost_sr,
        'total_sale_sr': total_sale_sr,
        'total_profit_sr': total_profit_sr,
    })


@require_interface_access('can_access_reports')
def showroom_performance_report(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    tenant_alias = _require_tenant_alias_or_raise('لا توجد بيئة معرض مفعلة لاستخراج تقرير أداء المعرض.')
    performance = ReportService.showroom_performance(tenant_alias=tenant_alias)

    sold_sales = list(
        Sale.objects.using(tenant_alias)
        .select_related('car', 'customer')
        .order_by('-sale_date', '-id')
    )

    total_sales_sr = Decimal('0')
    total_profit_sr = Decimal('0')
    for sale in sold_sales:
        currency = getattr(sale.car, 'currency', 'SR')
        total_sales_sr += _to_sr(sale.sale_price, currency)
        total_profit_sr += _to_sr(sale.actual_profit, currency)

    sold_count = len(sold_sales)
    average_profit_sr = (total_profit_sr / sold_count) if sold_count else Decimal('0')
    available_count = Car.objects.using(tenant_alias).filter(is_sold=False).count()
    turnover_ratio = (Decimal(sold_count) / Decimal(available_count)) if available_count else Decimal(sold_count)

    return render(request, 'sales/showroom_performance_report.html', {
        'performance': performance,
        'total_sales_sr': total_sales_sr,
        'sold_count': sold_count,
        'average_profit_sr': average_profit_sr,
        'average_days_in_inventory': performance.get('average_days_in_inventory', 0),
        'available_count': available_count,
        'turnover_ratio': turnover_ratio,
    })


@require_interface_access('can_access_reports')
def inventory_turnover_report(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    tenant_alias = _require_tenant_alias_or_raise('لا توجد بيئة معرض مفعلة لاستخراج تقرير دوران المخزون.')
    today = timezone.localdate()

    available_count = Car.objects.using(tenant_alias).filter(is_sold=False).count()
    total_sold_count = Sale.objects.using(tenant_alias).count()
    turnover_ratio = ReportService.inventory_turnover(tenant_alias=tenant_alias)

    sold_30_days = Sale.objects.using(tenant_alias).filter(sale_date__date__gte=today - timedelta(days=30)).count()
    sold_60_days = Sale.objects.using(tenant_alias).filter(sale_date__date__gte=today - timedelta(days=60)).count()
    sold_90_days = Sale.objects.using(tenant_alias).filter(sale_date__date__gte=today - timedelta(days=90)).count()

    movements = list(ReportService.inventory_movements(tenant_alias=tenant_alias)[:120])

    return render(request, 'sales/inventory_turnover_report.html', {
        'today': today,
        'available_count': available_count,
        'total_sold_count': total_sold_count,
        'turnover_ratio': turnover_ratio,
        'sold_30_days': sold_30_days,
        'sold_60_days': sold_60_days,
        'sold_90_days': sold_90_days,
        'movements': movements,
    })


@require_interface_access('can_access_reports')
def stale_cars_report(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    tenant_alias = _require_tenant_alias_or_raise('لا توجد بيئة معرض مفعلة لاستخراج تقرير السيارات الراكدة.')
    today = timezone.localdate()

    stale_rows = []
    stale_30 = 0
    stale_60 = 0
    stale_90 = 0

    cars = Car.objects.using(tenant_alias).filter(is_sold=False).order_by('created_at')
    for car in cars:
        age_days = (today - car.created_at.date()).days if car.created_at else 0
        if age_days < 30:
            continue

        if age_days >= 90:
            aging_bucket = '90+'
            stale_90 += 1
            stale_60 += 1
            stale_30 += 1
        elif age_days >= 60:
            aging_bucket = '60+'
            stale_60 += 1
            stale_30 += 1
        else:
            aging_bucket = '30+'
            stale_30 += 1

        stale_rows.append({
            'car': car,
            'age_days': age_days,
            'aging_bucket': aging_bucket,
            'total_cost': car.total_cost_price,
        })

    stale_rows.sort(key=lambda row: row['age_days'], reverse=True)

    return render(request, 'sales/stale_cars_report.html', {
        'today': today,
        'rows': stale_rows,
        'stale_30': stale_30,
        'stale_60': stale_60,
        'stale_90': stale_90,
    })


@require_interface_access('can_access_reports')
def daily_closing_control(request):
    tenant_alias = (get_current_tenant_db_alias() or '').strip()
    if not tenant_alias.startswith('tenant_'):
        raise PermissionDenied('لا توجد بيئة معرض مفعلة لإغلاق اليومية.')

    if not (request.user.is_staff or request.user.is_superuser or _is_platform_owner_session(request)):
        raise PermissionDenied('يلزم صلاحية إدارية لإغلاق اليومية.')

    today = timezone.localdate()
    closings_qs = DailyClosing.objects.using(tenant_alias).select_related('closed_by').order_by('-closing_date', '-created_at')
    last_closing = closings_qs.first()

    if request.method == 'POST':
        form = DailyClosingForm(request.POST)
        if form.is_valid():
            closing_date = form.cleaned_data['closing_date']
            if closing_date > today:
                form.add_error('closing_date', 'لا يمكن إغلاق يومية بتاريخ مستقبلي.')
            elif DailyClosing.objects.using(tenant_alias).filter(closing_date=closing_date).exists():
                form.add_error('closing_date', 'هذه اليومية مغلقة مسبقًا.')
            else:
                closing = form.save(commit=False)
                closing.closed_by_id = request.user.pk
                closing.save(using=tenant_alias)
                return redirect(f"{reverse('daily_closing')}?saved=1")
    else:
        form = DailyClosingForm(initial={'closing_date': today})

    return render(request, 'sales/daily_closing.html', {
        'form': form,
        'today': today,
        'last_closing': last_closing,
        'is_today_closed': closings_qs.filter(closing_date=today).exists(),
        'closings_count': closings_qs.count(),
        'closings': closings_qs[:30],
        'daily_closing_ok_url': reverse('daily_closing'),
    })


def _resolve_month_window(reference_date):
    month_start = reference_date.replace(day=1)
    if month_start.month == 12:
        next_month_start = month_start.replace(year=month_start.year + 1, month=1, day=1)
    else:
        next_month_start = month_start.replace(month=month_start.month + 1, day=1)
    month_end = next_month_start - timedelta(days=1)
    return month_start, month_end


def _backfill_voucher_journal_entries(alias):
    vouchers = FinanceVoucher.objects.using(alias).all().order_by('id')
    for voucher in vouchers.iterator(chunk_size=250):
        sync_journal_entry_for_voucher(voucher, alias=alias)


@require_interface_access('can_access_reports')
def chart_of_accounts(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    tenant_alias = _require_tenant_alias_or_raise('لا توجد بيئة معرض مفعلة لإدارة شجرة الحسابات.')
    ensure_default_chart_of_accounts(alias=tenant_alias)

    if request.method == 'POST':
        form = FinancialAccountForm(request.POST, tenant_alias=tenant_alias)
        if form.is_valid():
            account = form.save(commit=False)
            account.save(using=tenant_alias)
            return redirect(f"{reverse('chart_of_accounts')}?saved=1")
    else:
        form = FinancialAccountForm(tenant_alias=tenant_alias)

    accounts = list(
        FinancialAccount.objects.using(tenant_alias)
        .select_related('parent')
        .order_by('code', 'id')
    )

    return render(request, 'sales/chart_of_accounts.html', {
        'form': form,
        'accounts': accounts,
        'chart_ok_url': reverse('chart_of_accounts'),
    })


@require_interface_access('can_access_reports')
def financial_containers_management(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    tenant_alias = _require_tenant_alias_or_raise('لا توجد بيئة معرض مفعلة لإدارة الأوعية المالية.')
    ensure_default_chart_of_accounts(alias=tenant_alias)
    ensure_default_financial_containers(alias=tenant_alias)

    if request.method == 'POST':
        form = FinancialContainerForm(request.POST, tenant_alias=tenant_alias)
        if form.is_valid():
            container = form.save(commit=False)
            container.save(using=tenant_alias)
            return redirect(f"{reverse('financial_containers')}?saved=1")
    else:
        form = FinancialContainerForm(tenant_alias=tenant_alias)

    containers = list(
        FinancialContainer.objects.using(tenant_alias)
        .select_related('linked_account')
        .order_by('name', 'id')
    )

    return render(request, 'sales/financial_containers.html', {
        'form': form,
        'containers': containers,
        'containers_ok_url': reverse('financial_containers'),
    })


@require_interface_access('can_access_reports')
def trial_balance_report(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    tenant_alias = _require_tenant_alias_or_raise('لا توجد بيئة معرض مفعلة لاستخراج ميزان المراجعة.')
    ensure_default_chart_of_accounts(alias=tenant_alias)
    _backfill_voucher_journal_entries(tenant_alias)

    as_of_text = (request.GET.get('as_of_date') or '').strip()
    as_of_date = timezone.localdate()
    if as_of_text:
        try:
            as_of_date = datetime.strptime(as_of_text, '%Y-%m-%d').date()
        except ValueError:
            as_of_date = timezone.localdate()

    trial_data = build_trial_balance_rows(alias=tenant_alias, as_of_date=as_of_date)

    return render(request, 'sales/trial_balance_report.html', {
        'as_of_date': as_of_date,
        'rows': trial_data['rows'],
        'total_debit': trial_data['total_debit'],
        'total_credit': trial_data['total_credit'],
        'difference': trial_data['difference'],
        'is_balanced': trial_data['difference'] == Decimal('0'),
    })


@require_interface_access('can_access_reports')
def fiscal_period_closing_control(request):
    tenant_alias = _require_tenant_alias_or_raise('لا توجد بيئة معرض مفعلة لإدارة الإغلاق الشهري.')

    if not (request.user.is_staff or request.user.is_superuser or _is_platform_owner_session(request)):
        raise PermissionDenied('يلزم صلاحية إدارية لإدارة الإغلاق الشهري.')

    today = timezone.localdate()
    close_form = FiscalPeriodClosingForm()
    unlock_form = FiscalUnlockRequestForm(tenant_alias=tenant_alias)

    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()

        if action == 'close_month':
            close_form = FiscalPeriodClosingForm(request.POST)
            unlock_form = FiscalUnlockRequestForm(tenant_alias=tenant_alias)
            if close_form.is_valid():
                reference_date = close_form.cleaned_data['reference_date']
                month_start, month_end = _resolve_month_window(reference_date)
                current_month_start = today.replace(day=1)

                if month_start >= current_month_start:
                    close_form.add_error('reference_date', 'يمكن إغلاق الأشهر المكتملة فقط (الشهر الحالي غير مكتمل).')
                elif FiscalPeriodClosing.objects.using(tenant_alias).filter(
                    period_type=FiscalPeriodClosing.PERIOD_MONTHLY,
                    period_start=month_start,
                    period_end=month_end,
                ).exists():
                    close_form.add_error('reference_date', 'هذه الفترة الشهرية مغلقة مسبقًا.')
                else:
                    FiscalPeriodClosing.objects.using(tenant_alias).create(
                        period_type=FiscalPeriodClosing.PERIOD_MONTHLY,
                        period_start=month_start,
                        period_end=month_end,
                        is_locked=True,
                        closed_by_id=request.user.pk,
                        notes=close_form.cleaned_data.get('notes', ''),
                    )
                    write_platform_audit(
                        event_type='fiscal_period_close',
                        tenant_id=request.session.get('tenant_id') or '',
                        actor_username=request.user.username,
                        notes=f'إغلاق شهري للفترة {month_start} -> {month_end}',
                    )
                    return redirect(f"{reverse('fiscal_period_closing')}?closed=1")

        elif action == 'request_unlock':
            close_form = FiscalPeriodClosingForm()
            unlock_form = FiscalUnlockRequestForm(request.POST, tenant_alias=tenant_alias)
            if unlock_form.is_valid():
                closing = unlock_form.cleaned_data['closing']
                has_pending = FiscalUnlockRequest.objects.using(tenant_alias).filter(
                    closing_id=closing.pk,
                    status=FiscalUnlockRequest.STATUS_PENDING,
                ).exists()
                if has_pending:
                    unlock_form.add_error('closing', 'يوجد طلب فك إغلاق قيد المراجعة لهذه الفترة.')
                else:
                    FiscalUnlockRequest.objects.using(tenant_alias).create(
                        closing_id=closing.pk,
                        requested_by_id=request.user.pk,
                        reason=unlock_form.cleaned_data['reason'],
                        status=FiscalUnlockRequest.STATUS_PENDING,
                    )
                    write_platform_audit(
                        event_type='fiscal_unlock_request',
                        tenant_id=request.session.get('tenant_id') or '',
                        actor_username=request.user.username,
                        notes=f'طلب فك إغلاق للفترة {closing.period_start} -> {closing.period_end}',
                    )
                    return redirect(f"{reverse('fiscal_period_closing')}?unlock_requested=1")

        elif action == 'review_unlock':
            if not request.user.is_superuser:
                raise PermissionDenied('فقط المشرف العام يمكنه مراجعة طلبات فك الإغلاق.')

            unlock_request_id = (request.POST.get('unlock_request_id') or '').strip()
            decision = (request.POST.get('decision') or '').strip()
            review_notes = (request.POST.get('review_notes') or '').strip()[:255]

            unlock_request = (
                FiscalUnlockRequest.objects.using(tenant_alias)
                .select_related('closing')
                .filter(pk=unlock_request_id)
                .first()
            )

            if unlock_request and unlock_request.status == FiscalUnlockRequest.STATUS_PENDING:
                audit_event = 'fiscal_unlock_rejected'
                if decision == 'approve':
                    unlock_request.status = FiscalUnlockRequest.STATUS_APPROVED
                    unlock_request.closing.is_locked = False
                    unlock_request.closing.save(using=tenant_alias, update_fields=['is_locked'])
                    audit_event = 'fiscal_unlock_approved'
                else:
                    unlock_request.status = FiscalUnlockRequest.STATUS_REJECTED

                unlock_request.reviewed_by_id = request.user.pk
                unlock_request.review_notes = review_notes
                unlock_request.reviewed_at = timezone.now()
                unlock_request.save(
                    using=tenant_alias,
                    update_fields=['status', 'reviewed_by', 'review_notes', 'reviewed_at'],
                )

                write_platform_audit(
                    event_type=audit_event,
                    tenant_id=request.session.get('tenant_id') or '',
                    actor_username=request.user.username,
                    notes=(
                        f"مراجعة طلب فك الإغلاق للفترة "
                        f"{unlock_request.closing.period_start} -> {unlock_request.closing.period_end}"
                    ),
                )
                return redirect(f"{reverse('fiscal_period_closing')}?unlock_reviewed=1")

    closings = list(
        FiscalPeriodClosing.objects.using(tenant_alias)
        .select_related('closed_by')
        .order_by('-period_end', '-created_at')[:24]
    )
    unlock_requests = list(
        FiscalUnlockRequest.objects.using(tenant_alias)
        .select_related('closing', 'requested_by', 'reviewed_by')
        .order_by('-created_at')[:40]
    )

    return render(request, 'sales/fiscal_period_closing.html', {
        'today': today,
        'close_form': close_form,
        'unlock_form': unlock_form,
        'closings': closings,
        'unlock_requests': unlock_requests,
        'fiscal_closing_ok_url': reverse('fiscal_period_closing'),
    })


def _sum_converted_values(queryset, amount_field, currency_field):
    total = Decimal('0')
    for amount, currency in queryset.values_list(amount_field, currency_field).iterator(chunk_size=300):
        total += _to_sr(amount, currency)
    return total


def _database_engine_label():
    engine = (settings.DATABASES.get('default', {}).get('ENGINE') or '').lower()
    if 'postgresql' in engine:
        return 'PostgreSQL'
    if 'sqlite' in engine:
        return 'SQLite'
    if 'mysql' in engine:
        return 'MySQL'
    return (engine.split('.')[-1] or 'Unknown').upper()


def _application_server_label(request):
    server_software = (request.META.get('SERVER_SOFTWARE') or '').strip()
    normalized = server_software.lower()
    if 'gunicorn' in normalized:
        return 'Gunicorn'
    if 'uvicorn' in normalized:
        return 'Uvicorn'
    if 'wsgi' in normalized or 'runserver' in normalized:
        return server_software or 'Django WSGI'
    return server_software or 'Application Server'


def _proxy_status(request):
    forwarded_headers = [
        request.META.get('HTTP_X_FORWARDED_FOR'),
        request.META.get('HTTP_X_FORWARDED_PROTO'),
        request.META.get('HTTP_X_REAL_IP'),
    ]
    has_proxy = any(bool(value) for value in forwarded_headers)
    if has_proxy:
        return True, 'Nginx/Reverse Proxy مكتشف'
    return False, 'لم يتم اكتشاف Proxy في هذه الجلسة'


def _collect_active_tenant_sessions():
    active_sessions = Session.objects.filter(expire_date__gte=timezone.now())
    active_connection_count = 0
    connected_tenants = set()
    decode_errors = 0

    for session in active_sessions:
        try:
            payload = session.get_decoded()
        except Exception:
            decode_errors += 1
            continue

        tenant_id = normalize_tenant_id(payload.get('tenant_id'))
        tenant_alias = (payload.get(TENANT_DB_ALIAS_SESSION_KEY) or '').strip()
        if not tenant_id:
            continue

        if tenant_alias == f'tenant_{tenant_id}':
            active_connection_count += 1
            connected_tenants.add(tenant_id)

    return active_connection_count, len(connected_tenants), decode_errors


def _collect_tenant_active_sessions(target_tenant_id):
    normalized_tenant = normalize_tenant_id(target_tenant_id)
    if not normalized_tenant:
        return 0, 0, 0

    active_sessions = Session.objects.filter(expire_date__gte=timezone.now())
    active_connection_count = 0
    decode_errors = 0
    expected_alias = f'tenant_{normalized_tenant}'

    for session in active_sessions:
        try:
            payload = session.get_decoded()
        except Exception:
            decode_errors += 1
            continue

        tenant_id = normalize_tenant_id(payload.get('tenant_id'))
        tenant_alias = (payload.get(TENANT_DB_ALIAS_SESSION_KEY) or '').strip()
        if tenant_id == normalized_tenant and tenant_alias == expected_alias:
            active_connection_count += 1

    connected_tenants = 1 if active_connection_count > 0 else 0
    return active_connection_count, connected_tenants, decode_errors


def _default_business_rows_snapshot():
    rows = []
    total = 0
    try:
        default_tables = set(connections['default'].introspection.table_names())
    except Exception:
        default_tables = set()

    for model in CENTRAL_TENANT_BUSINESS_MODELS:
        table_name = model._meta.db_table
        if table_name not in default_tables:
            rows.append({
                'table': table_name,
                'model': model.__name__,
                'row_count': 0,
                'table_exists': False,
            })
            continue

        count = model.objects.using('default').count()
        total += count
        rows.append({
            'table': table_name,
            'model': model.__name__,
            'row_count': count,
            'table_exists': True,
        })
    return total, rows


def _build_tenant_runtime_snapshot(active_tenants):
    snapshot = {
        'healthy_tenants_count': 0,
        'tenant_connection_failures': [],
        'tenant_query_failures': [],
        'tenant_schema_issues': [],
        'alias_map': {},
        'total_sales_sr': Decimal('0'),
        'operating_expenses_sr': Decimal('0'),
        'general_expenses_sr': Decimal('0'),
        'maintenance_expenses_sr': Decimal('0'),
        'voucher_outflows_sr': Decimal('0'),
    }

    for tenant in active_tenants:
        tenant_id = normalize_tenant_id(tenant.tenant_id)
        alias = ensure_tenant_connection(tenant_id)
        if not alias:
            snapshot['tenant_connection_failures'].append(f'تعذر إنشاء اتصال قاعدة بيانات للمعرض: {tenant_id}')
            continue

        snapshot['alias_map'][tenant_id] = alias

        try:
            with connections[alias].cursor() as cursor:
                cursor.execute('SELECT 1')
            table_names = set(connections[alias].introspection.table_names())
        except Exception as exc:
            snapshot['tenant_connection_failures'].append(
                f'فشل الاتصال بقاعدة المعرض {tenant_id}: {str(exc)[:140]}'
            )
            continue

        snapshot['healthy_tenants_count'] += 1

        if 'sales_platformtenant' in table_names:
            snapshot['tenant_schema_issues'].append(
                f'فشل عزل: جدول المنصة موجود داخل قاعدة المعرض {tenant_id}.'
            )

        try:
            snapshot['total_sales_sr'] += _sum_converted_values(
                Sale.objects.using(alias).all(),
                'sale_price',
                'car__currency',
            )
            snapshot['operating_expenses_sr'] += Expense.objects.using(alias).aggregate(total=Sum('amount'))['total'] or Decimal('0')
            snapshot['general_expenses_sr'] += _sum_converted_values(
                GeneralExpense.objects.using(alias).all(),
                'amount',
                'currency',
            )
            snapshot['maintenance_expenses_sr'] += _sum_converted_values(
                CarMaintenance.objects.using(alias).all(),
                'amount',
                'car__currency',
            )
            snapshot['voucher_outflows_sr'] += _sum_converted_values(
                FinanceVoucher.objects.using(alias).filter(voucher_type__in=['operating', 'payment']),
                'amount',
                'currency',
            )
        except Exception as exc:
            snapshot['tenant_query_failures'].append(
                f'فشل قراءة مؤشرات المعرض {tenant_id}: {str(exc)[:140]}'
            )

    snapshot['total_expenses_sr'] = (
        snapshot['operating_expenses_sr']
        + snapshot['general_expenses_sr']
        + snapshot['maintenance_expenses_sr']
        + snapshot['voucher_outflows_sr']
    )
    snapshot['current_balance_sr'] = snapshot['total_sales_sr'] - snapshot['total_expenses_sr']
    return snapshot


def _safe_storage_file_size(file_name):
    if not file_name:
        return 0

    try:
        return int(default_storage.size(file_name))
    except Exception:
        return 0


def _sum_file_field_storage_bytes(queryset, field_name):
    total = 0
    for file_name in queryset.exclude(**{field_name: ''}).values_list(field_name, flat=True).iterator(chunk_size=250):
        total += _safe_storage_file_size(file_name)
    return total


def _tenant_storage_usage_bytes(alias):
    total_bytes = 0
    total_bytes += _sum_file_field_storage_bytes(Car.objects.using(alias), 'image')
    total_bytes += _sum_file_field_storage_bytes(Car.objects.using(alias), 'contract_image')
    total_bytes += _sum_file_field_storage_bytes(Sale.objects.using(alias), 'sale_contract_image')
    total_bytes += _sum_file_field_storage_bytes(CarMaintenance.objects.using(alias), 'invoice_image')
    total_bytes += _sum_file_field_storage_bytes(CarDocument.objects.using(alias), 'file')
    return total_bytes


def _build_tenant_activity_heatmap(active_tenants, alias_map):
    heat_rows = []
    recent_cutoff = timezone.now() - timedelta(days=30)

    for tenant in active_tenants:
        tenant_id = normalize_tenant_id(tenant.tenant_id)
        alias = alias_map.get(tenant_id) or ensure_tenant_connection(tenant_id)
        if not alias:
            heat_rows.append({
                'tenant_id': tenant_id,
                'tenant_name': tenant.name,
                'users_count': 0,
                'cars_count': 0,
                'sold_cars_count': 0,
                'vouchers_count': 0,
                'operations_30d': 0,
                'storage_used_mb': 0,
                'storage_quota_mb': tenant.max_storage_mb,
                'storage_usage_pct': 0,
                'activity_score': 0,
                'heat_level': 'na',
                'heat_label': 'غير متاح',
                'heat_class': 'heat-na',
            })
            continue

        try:
            users_count = User.objects.using(alias).count()
            cars_count = Car.objects.using(alias).count()
            sold_cars_count = Car.objects.using(alias).filter(is_sold=True).count()
            vouchers_count = FinanceVoucher.objects.using(alias).count()
            operations_30d = AuditLog.objects.using(alias).filter(timestamp__gte=recent_cutoff).count()
            storage_bytes = _tenant_storage_usage_bytes(alias)
        except Exception:
            heat_rows.append({
                'tenant_id': tenant_id,
                'tenant_name': tenant.name,
                'users_count': 0,
                'cars_count': 0,
                'sold_cars_count': 0,
                'vouchers_count': 0,
                'operations_30d': 0,
                'storage_used_mb': 0,
                'storage_quota_mb': tenant.max_storage_mb,
                'storage_usage_pct': 0,
                'activity_score': 0,
                'heat_level': 'na',
                'heat_label': 'خطأ قراءة',
                'heat_class': 'heat-na',
            })
            continue

        storage_used_mb = Decimal(storage_bytes) / Decimal(1024 * 1024)
        quota_mb = Decimal(tenant.max_storage_mb or 0)
        if quota_mb > 0:
            storage_usage_pct = float((storage_used_mb / quota_mb) * Decimal('100'))
        else:
            storage_usage_pct = 0.0

        activity_score = min(
            100,
            int((operations_30d * 2) + (vouchers_count * 0.4) + (sold_cars_count * 2) + (users_count * 4)),
        )

        if storage_usage_pct >= 85 or activity_score >= 75:
            heat_level = 'high'
            heat_label = 'مرتفع'
            heat_class = 'heat-high'
        elif storage_usage_pct >= 55 or activity_score >= 40:
            heat_level = 'medium'
            heat_label = 'متوسط'
            heat_class = 'heat-medium'
        else:
            heat_level = 'low'
            heat_label = 'منخفض'
            heat_class = 'heat-low'

        heat_rows.append({
            'tenant_id': tenant_id,
            'tenant_name': tenant.name,
            'users_count': users_count,
            'cars_count': cars_count,
            'sold_cars_count': sold_cars_count,
            'vouchers_count': vouchers_count,
            'operations_30d': operations_30d,
            'storage_used_mb': float(storage_used_mb),
            'storage_quota_mb': tenant.max_storage_mb,
            'storage_usage_pct': storage_usage_pct,
            'activity_score': activity_score,
            'heat_level': heat_level,
            'heat_label': heat_label,
            'heat_class': heat_class,
        })

    heat_rank = {'high': 3, 'medium': 2, 'low': 1, 'na': 0}
    heat_rows.sort(
        key=lambda item: (
            heat_rank.get(item.get('heat_level'), 0),
            item.get('activity_score', 0),
            item.get('storage_usage_pct', 0),
        ),
        reverse=True,
    )
    return heat_rows


def _collect_central_activity_logs(active_tenants, alias_map, include_global_logs=True):
    events = []
    errors = []

    if include_global_logs:
        global_logs = GlobalAuditLog.objects.using('default').order_by('-created_at')[:60]
        for log in global_logs:
            notes = (log.notes or '').strip()
            status = 'failed' if ('failed' in (log.event_type or '').lower() or 'فشل' in notes) else 'success'
            operation_text = f"{log.get_event_type_display()} - {notes or 'بدون ملاحظات'}"
            events.append({
                'timestamp': log.created_at,
                'showroom_id': log.tenant_id or 'platform',
                'operation': operation_text,
                'status': status,
            })

    action_label_map = {
        'create': 'إضافة',
        'update': 'تعديل',
        'delete': 'حذف',
    }

    for tenant in active_tenants:
        tenant_id = normalize_tenant_id(tenant.tenant_id)
        alias = alias_map.get(tenant_id) or ensure_tenant_connection(tenant_id)
        if not alias:
            continue

        try:
            tenant_logs = AuditLog.objects.using(alias).select_related('user').order_by('-timestamp')[:CENTRAL_AUDIT_TENANT_LOG_LIMIT]
            for log in tenant_logs:
                raw_action = (log.action or '').strip().lower()
                action_label = action_label_map.get(raw_action, (log.action or 'عملية').upper())
                status = 'failed' if raw_action in {'failed', 'error', 'failure'} else 'success'
                events.append({
                    'timestamp': log.timestamp,
                    'showroom_id': tenant_id,
                    'operation': f"{action_label}: {log.target_model or '-'} ({log.target_pk or '-'})",
                    'status': status,
                })
        except Exception as exc:
            errors.append(f'تعذر قراءة سجل التدقيق من المعرض {tenant_id}: {str(exc)[:120]}')
            events.append({
                'timestamp': timezone.now(),
                'showroom_id': tenant_id,
                'operation': 'فشل قراءة سجل التدقيق لهذا المعرض.',
                'status': 'failed',
            })

    events.sort(key=lambda item: item.get('timestamp') or timezone.now(), reverse=True)
    return events[:CENTRAL_AUDIT_EVENT_LIMIT], errors


def _parse_platform_login_failure_note(raw_note):
    note = (raw_note or '').strip()
    parts = [part.strip() for part in note.split('|') if part.strip()]

    reason = parts[0] if parts else 'unknown'
    ip_address = ''
    user_agent = ''

    for part in parts[1:]:
        if part.startswith('ip='):
            ip_address = part[3:].strip()
        elif part.startswith('ua='):
            user_agent = part[3:].strip()

    return {
        'reason': reason,
        'ip_address': ip_address,
        'user_agent': user_agent,
    }


def _platform_failed_login_logs_snapshot(query_params):
    username_q = (query_params.get('platform_user_q') or '').strip()
    ip_q = (query_params.get('platform_ip_q') or '').strip()
    start_date = (query_params.get('platform_start_date') or '').strip()
    end_date = (query_params.get('platform_end_date') or '').strip()

    logs = GlobalAuditLog.objects.using('default').filter(event_type='platform_login_failed')

    if username_q:
        logs = logs.filter(actor_username__icontains=username_q)
    if ip_q:
        logs = logs.filter(notes__icontains=f"ip={ip_q}")
    if start_date:
        logs = logs.filter(created_at__date__gte=start_date)
    if end_date:
        logs = logs.filter(created_at__date__lte=end_date)

    rows = []
    for log in logs.order_by('-created_at')[:120]:
        parsed = _parse_platform_login_failure_note(log.notes)
        rows.append({
            'timestamp': log.created_at,
            'username': log.actor_username or '-',
            'reason': parsed['reason'] or '-',
            'ip_address': parsed['ip_address'] or '-',
            'user_agent': parsed['user_agent'] or '-',
        })

    return rows, {
        'platform_user_q': username_q,
        'platform_ip_q': ip_q,
        'platform_start_date': start_date,
        'platform_end_date': end_date,
    }


def _manual_trace_lookup(vin_query, voucher_query, active_tenants, alias_map):
    results = []
    errors = []
    vin_value = (vin_query or '').strip().upper()
    voucher_value = (voucher_query or '').strip().upper()

    if not vin_value and not voucher_value:
        return results, errors

    return _manual_trace_lookup_scoped(
        vin_value,
        voucher_value,
        active_tenants,
        alias_map,
        include_default_checks=True,
    )


def _manual_trace_lookup_scoped(vin_value, voucher_value, active_tenants, alias_map, include_default_checks):
    results = []
    errors = []

    if not vin_value and not voucher_value:
        return results, errors

    if include_default_checks:
        try:
            if vin_value:
                leaked_default_cars = list(
                    Car.objects.using('default').filter(vin__icontains=vin_value).order_by('-created_at')[:5]
                )
                for car in leaked_default_cars:
                    results.append({
                        'timestamp': car.created_at,
                        'showroom_id': 'default',
                        'lookup_type': 'VIN',
                        'operation': f'تم العثور على السيارة {car.vin} داخل قاعدة المنصة (حالة غير طبيعية).',
                        'status': 'failed',
                    })

            if voucher_value:
                leaked_default_vouchers = list(
                    FinanceVoucher.objects.using('default')
                    .filter(voucher_number__icontains=voucher_value)
                    .order_by('-created_at')[:5]
                )
                for voucher in leaked_default_vouchers:
                    results.append({
                        'timestamp': voucher.created_at,
                        'showroom_id': 'default',
                        'lookup_type': 'فاتورة',
                        'operation': f'تم العثور على الفاتورة {voucher.voucher_number} داخل قاعدة المنصة (حالة غير طبيعية).',
                        'status': 'failed',
                    })
        except Exception as exc:
            errors.append(f'تعذر فحص قاعدة المنصة أثناء البحث اليدوي: {str(exc)[:120]}')

    for tenant in active_tenants:
        tenant_id = normalize_tenant_id(tenant.tenant_id)
        alias = alias_map.get(tenant_id) or ensure_tenant_connection(tenant_id)
        if not alias:
            continue

        if vin_value:
            try:
                cars = list(Car.objects.using(alias).filter(vin__icontains=vin_value).order_by('-created_at')[:5])
                for car in cars:
                    sale = Sale.objects.using(alias).filter(car_id=car.id).order_by('-sale_date').first()
                    if sale:
                        operation = f'VIN مطابق: {car.vin} - بيع #{sale.id} بتاريخ {sale.sale_date:%Y-%m-%d}.'
                        timestamp = sale.sale_date
                    else:
                        operation = f'VIN مطابق: {car.vin} - السيارة متاحة ولم يتم بيعها بعد.'
                        timestamp = car.created_at

                    results.append({
                        'timestamp': timestamp,
                        'showroom_id': tenant_id,
                        'lookup_type': 'VIN',
                        'operation': operation,
                        'status': 'success',
                    })
            except Exception as exc:
                errors.append(f'فشل بحث VIN داخل معرض {tenant_id}: {str(exc)[:120]}')

        if voucher_value:
            try:
                vouchers = list(
                    FinanceVoucher.objects.using(alias)
                    .filter(voucher_number__icontains=voucher_value)
                    .order_by('-created_at')[:5]
                )
                for voucher in vouchers:
                    results.append({
                        'timestamp': voucher.created_at,
                        'showroom_id': tenant_id,
                        'lookup_type': 'فاتورة',
                        'operation': (
                            f"فاتورة مطابقة: {voucher.voucher_number} "
                            f"({voucher.get_voucher_type_display()}) بقيمة {voucher.amount} {voucher.currency}."
                        ),
                        'status': 'success',
                    })
            except Exception as exc:
                errors.append(f'فشل بحث الفاتورة داخل معرض {tenant_id}: {str(exc)[:120]}')

    results.sort(key=lambda item: item.get('timestamp') or timezone.now(), reverse=True)
    return results[:120], errors


@login_required
def central_audit_monitor(request):
    can_view_global_scope = _has_platform_wide_monitor_access(request)
    tenant_id = normalize_tenant_id(request.session.get('tenant_id'))
    tenant_alias = (request.session.get(TENANT_DB_ALIAS_SESSION_KEY) or '').strip()

    if not can_view_global_scope:
        if not tenant_id or tenant_alias != f'tenant_{tenant_id}':
            raise PermissionDenied('يمكنك الاطلاع على بيانات معرضك فقط.')

        if not request.user.is_superuser:
            access, _ = InterfaceAccess.objects.using(tenant_alias).get_or_create(user_id=request.user.pk)
            if not access.can_access_reports:
                raise PermissionDenied('ليس لديك صلاحية الوصول إلى هذه الصفحة.')

    technical_errors = []

    if can_view_global_scope:
        active_tenants = list(
            PlatformTenant.objects.using('default').filter(is_active=True, is_deleted=False).order_by('tenant_id')
        )

        database_ok = True
        try:
            with connections['default'].cursor() as cursor:
                cursor.execute('SELECT 1')
        except Exception as exc:
            database_ok = False
            technical_errors.append(f'فشل الاتصال بقاعدة بيانات المنصة: {str(exc)[:140]}')

        active_connection_count, connected_tenants_count, session_decode_errors = _collect_active_tenant_sessions()
        default_business_rows, default_business_details = _default_business_rows_snapshot()
        if default_business_rows > 0:
            technical_errors.append(
                f'فشل عزل: تم العثور على {default_business_rows} سجل أعمال داخل قاعدة المنصة (default).'
            )

        scope_label = 'نطاق العرض: جميع المعارض (صلاحية أدمن المنصة)'
        scope_showroom_id = 'all'
    else:
        tenant_record = PlatformTenant.objects.using('default').filter(
            tenant_id=tenant_id,
            is_active=True,
            is_deleted=False,
        ).first()
        if tenant_record is None:
            raise PermissionDenied('لا يمكن عرض هذه الصفحة لأن معرف المعرض غير صالح أو غير نشط.')

        active_tenants = [tenant_record]
        active_connection_count, connected_tenants_count, session_decode_errors = _collect_tenant_active_sessions(tenant_id)

        database_ok = True
        try:
            with connections[tenant_alias].cursor() as cursor:
                cursor.execute('SELECT 1')
        except Exception as exc:
            database_ok = False
            technical_errors.append(f'فشل الاتصال بقاعدة بيانات المعرض {tenant_id}: {str(exc)[:140]}')

        if connected_tenants_count == 0:
            connected_tenants_count = 1

        default_business_rows = 0
        default_business_details = []
        scope_label = f'نطاق العرض: بيانات معرضك فقط ({tenant_id})'
        scope_showroom_id = tenant_id

    if session_decode_errors:
        technical_errors.append(f'تم تجاهل {session_decode_errors} جلسة تالفة أثناء تحليل الاتصالات النشطة.')

    snapshot = _build_tenant_runtime_snapshot(active_tenants)
    technical_errors.extend(snapshot['tenant_connection_failures'])
    technical_errors.extend(snapshot['tenant_query_failures'])
    technical_errors.extend(snapshot['tenant_schema_issues'])
    tenant_activity_rows = _build_tenant_activity_heatmap(active_tenants, snapshot['alias_map'])

    activity_events, activity_errors = _collect_central_activity_logs(
        active_tenants,
        snapshot['alias_map'],
        include_global_logs=can_view_global_scope,
    )
    technical_errors.extend(activity_errors)

    platform_failed_login_rows = []
    platform_failed_login_filters = {
        'platform_user_q': '',
        'platform_ip_q': '',
        'platform_start_date': '',
        'platform_end_date': '',
    }
    if can_view_global_scope:
        platform_failed_login_rows, platform_failed_login_filters = _platform_failed_login_logs_snapshot(request.GET)

    vin_query = (request.GET.get('vin_query') or '').strip()
    voucher_query = (request.GET.get('voucher_query') or '').strip()
    trace_results, trace_errors = _manual_trace_lookup_scoped(
        (vin_query or '').strip().upper(),
        (voucher_query or '').strip().upper(),
        active_tenants,
        snapshot['alias_map'],
        include_default_checks=can_view_global_scope,
    )
    technical_errors.extend(trace_errors)

    deduped_errors = []
    seen_errors = set()
    for error_text in technical_errors:
        if not error_text:
            continue
        if error_text in seen_errors:
            continue
        seen_errors.add(error_text)
        deduped_errors.append(error_text)

    app_server_label = _application_server_label(request)
    proxy_ok, proxy_label = _proxy_status(request)
    db_engine_label = _database_engine_label()

    status_badges = [
        {
            'title': (
                f'قاعدة البيانات ({db_engine_label})'
                if can_view_global_scope
                else f'قاعدة بيانات المعرض ({db_engine_label})'
            ),
            'value': 'متصلة' if database_ok else 'غير متصلة',
            'state': 'ok' if database_ok else 'down',
            'icon': 'bi-database-fill-check' if database_ok else 'bi-database-x',
        },
        {
            'title': 'سيرفر التطبيق',
            'value': app_server_label,
            'state': 'ok',
            'icon': 'bi-hdd-network-fill',
        },
        {
            'title': 'Nginx / Proxy',
            'value': proxy_label,
            'state': 'ok' if proxy_ok else 'warn',
            'icon': 'bi-shield-check' if proxy_ok else 'bi-shield-exclamation',
        },
    ]

    financial_balance = snapshot['current_balance_sr']
    financial_integrity_ok = financial_balance >= Decimal('0')
    schema_isolation_issues_count = len(snapshot['tenant_schema_issues'])
    heat_high_count = sum(1 for row in tenant_activity_rows if row.get('heat_level') == 'high')
    heat_medium_count = sum(1 for row in tenant_activity_rows if row.get('heat_level') == 'medium')
    heat_low_count = sum(1 for row in tenant_activity_rows if row.get('heat_level') == 'low')
    if can_view_global_scope and default_business_rows > 0:
        schema_isolation_issues_count += 1

    return render(request, 'sales/central_audit_monitor.html', {
        'can_view_global_scope': can_view_global_scope,
        'scope_label': scope_label,
        'scope_showroom_id': scope_showroom_id,
        'status_badges': status_badges,
        'active_tenants_count': len(active_tenants),
        'healthy_tenants_count': snapshot['healthy_tenants_count'],
        'connected_tenants_count': connected_tenants_count,
        'active_connection_count': active_connection_count,
        'schema_isolation_issues_count': schema_isolation_issues_count,
        'default_business_rows': default_business_rows,
        'default_business_details': default_business_details,
        'total_sales_sr': snapshot['total_sales_sr'],
        'total_expenses_sr': snapshot['total_expenses_sr'],
        'financial_balance_sr': financial_balance,
        'financial_integrity_ok': financial_integrity_ok,
        'tenant_activity_rows': tenant_activity_rows,
        'heat_high_count': heat_high_count,
        'heat_medium_count': heat_medium_count,
        'heat_low_count': heat_low_count,
        'activity_events': activity_events,
        'platform_failed_login_rows': platform_failed_login_rows,
        'platform_failed_login_filters': platform_failed_login_filters,
        'vin_query': vin_query,
        'voucher_query': voucher_query,
        'trace_results': trace_results,
        'trace_has_query': bool(vin_query or voucher_query),
        'technical_errors': deduped_errors,
        'can_open_audit_report': request.user.is_superuser or can_view_global_scope,
    })


def _filtered_audit_logs_queryset(query_params):
    selected_user = (query_params.get('user') or '').strip()
    selected_model = (query_params.get('model') or '').strip()
    selected_action = (query_params.get('action') or '').strip()
    start_date = (query_params.get('start_date') or '').strip()
    end_date = (query_params.get('end_date') or '').strip()

    logs = AuditLog.objects.select_related('user').all()

    if selected_user.isdigit():
        logs = logs.filter(user_id=int(selected_user))
    if selected_model:
        logs = logs.filter(target_model=selected_model)
    if selected_action:
        logs = logs.filter(action=selected_action)
    if start_date:
        logs = logs.filter(timestamp__date__gte=start_date)
    if end_date:
        logs = logs.filter(timestamp__date__lte=end_date)

    return logs, {
        'selected_user': selected_user,
        'selected_model': selected_model,
        'selected_action': selected_action,
        'start_date': start_date,
        'end_date': end_date,
    }


def _build_audit_field_changes(before_data, after_data):
    if not isinstance(before_data, dict) or not isinstance(after_data, dict):
        return []

    changes = []
    for field_name in sorted(set(before_data.keys()) | set(after_data.keys())):
        old_value = before_data.get(field_name)
        new_value = after_data.get(field_name)
        if old_value == new_value:
            continue

        changes.append({
            'field': field_name,
            'before': old_value,
            'after': new_value,
        })

    return changes


@require_interface_access('can_access_reports')
def audit_logs_report(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    filtered_logs, filters_context = _filtered_audit_logs_queryset(request.GET)

    total_count = filtered_logs.count()
    logs = filtered_logs.order_by('-timestamp')[:250]

    for log in logs:
        log.field_changes = _build_audit_field_changes(log.before_data, log.after_data)

    users_with_logs = User.objects.filter(
        id__in=AuditLog.objects.exclude(user__isnull=True).values_list('user_id', flat=True).distinct()
    ).order_by('username')
    models_with_logs = [
        value
        for value in AuditLog.objects.order_by('target_model').values_list('target_model', flat=True).distinct()
        if value
    ]
    actions_with_logs = [
        value
        for value in AuditLog.objects.order_by('action').values_list('action', flat=True).distinct()
        if value
    ]

    return render(request, 'sales/audit_logs_report.html', {
        'logs': logs,
        'total_count': total_count,
        'users_with_logs': users_with_logs,
        'models_with_logs': models_with_logs,
        'actions_with_logs': actions_with_logs,
        **filters_context,
    })


@require_interface_access('can_access_reports')
def export_audit_logs_excel(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    logs, _ = _filtered_audit_logs_queryset(request.GET)
    logs = logs.order_by('-timestamp')

    reports_dir = Path(settings.MEDIA_ROOT) / 'reports'
    reports_dir.mkdir(parents=True, exist_ok=True)

    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    file_name = f'audit_logs_{timestamp}.xlsx'
    file_path = reports_dir / file_name

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'سجل التدقيق'
    sheet.append([
        'وقت العملية',
        'المستخدم',
        'الإجراء',
        'النموذج',
        'معرف السجل',
        'معرف المعرض',
        'المسار',
        'عنوان IP',
        'القيم السابقة',
        'القيم الجديدة',
    ])

    for log in logs:
        before_data = json.dumps(log.before_data, ensure_ascii=False, default=str) if log.before_data is not None else ''
        after_data = json.dumps(log.after_data, ensure_ascii=False, default=str) if log.after_data is not None else ''
        sheet.append([
            timezone.localtime(log.timestamp).strftime('%Y-%m-%d %H:%M:%S'),
            log.user.username if log.user else '',
            log.action,
            log.target_model,
            log.target_pk,
            log.tenant_id,
            log.request_path,
            log.ip_address,
            before_data,
            after_data,
        ])

    workbook.save(file_path)
    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=file_name)


@require_interface_access('can_access_reports')
def export_audit_logs_csv(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    logs, _ = _filtered_audit_logs_queryset(request.GET)
    logs = logs.order_by('-timestamp')

    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    file_name = f'audit_logs_{timestamp}.csv'

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow([
        'وقت العملية',
        'المستخدم',
        'الإجراء',
        'النموذج',
        'معرف السجل',
        'معرف المعرض',
        'المسار',
        'عنوان IP',
        'القيم السابقة',
        'القيم الجديدة',
    ])

    for log in logs:
        before_data = json.dumps(log.before_data, ensure_ascii=False, default=str) if log.before_data is not None else ''
        after_data = json.dumps(log.after_data, ensure_ascii=False, default=str) if log.after_data is not None else ''
        writer.writerow([
            timezone.localtime(log.timestamp).strftime('%Y-%m-%d %H:%M:%S'),
            log.user.username if log.user else '',
            log.action,
            log.target_model,
            log.target_pk,
            log.tenant_id,
            log.request_path,
            log.ip_address,
            before_data,
            after_data,
        ])

    return response


@require_interface_access('can_access_reports')
def general_expenses_management(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    form = GeneralExpenseForm(initial={'expense_date': timezone.localdate()})
    if request.method == 'POST':
        form = GeneralExpenseForm(request.POST)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.created_by = request.user
            expense.save()
            return redirect(f"{reverse('general_expenses_management')}?saved=1")

    expenses = GeneralExpense.objects.select_related('created_by').all().order_by('-expense_date', '-id')
    total_general_expenses = sum(
        (_to_sr(item.amount, item.currency) for item in expenses),
        Decimal('0'),
    )

    return render(request, 'sales/general_expenses.html', {
        'form': form,
        'expenses': expenses,
        'total_general_expenses': total_general_expenses,
    })


@require_interface_access('can_access_reports')
def digital_archive(request):
    if not request.user.is_staff:
        raise PermissionDenied

    today = timezone.localdate()
    warning_deadline = today + timedelta(days=30)
    cars = Car.objects.prefetch_related('documents').all().order_by('-created_at')

    archive_rows = []
    for car in cars:
        insurance_days_left = None
        registration_days_left = None

        if car.insurance_expiry:
            insurance_days_left = (car.insurance_expiry - today).days
        if car.registration_expiry:
            registration_days_left = (car.registration_expiry - today).days

        expiring_soon = (
            (car.insurance_expiry and today <= car.insurance_expiry <= warning_deadline)
            or (car.registration_expiry and today <= car.registration_expiry <= warning_deadline)
        )

        archive_rows.append({
            'car': car,
            'documents_count': car.documents.count(),
            'insurance_days_left': insurance_days_left,
            'registration_days_left': registration_days_left,
            'expiring_soon': expiring_soon,
        })

    return render(request, 'sales/digital_archive.html', {
        'archive_rows': archive_rows,
    })


@require_interface_access('can_access_cars')
def inventory_reconciliation(request):
    if not request.user.is_staff:
        raise PermissionDenied

    cars = list(Car.objects.filter(is_sold=False).order_by('vin'))
    report = None
    checked_ids = set()
    extra_vins_text = ''

    if request.method == 'POST':
        checked_ids = {
            int(raw_id)
            for raw_id in request.POST.getlist('checked_car_ids')
            if str(raw_id).isdigit()
        }
        extra_vins_text = request.POST.get('extra_vins', '')

        missing_cars = [car for car in cars if car.id not in checked_ids]
        normalized_system_vins = {car.vin.strip().upper() for car in cars}

        raw_extra_vins = [
            item.strip().upper()
            for item in re.split(r'[\n,]+', extra_vins_text)
            if item.strip()
        ]
        extra_not_in_system = sorted({vin for vin in raw_extra_vins if vin not in normalized_system_vins})

        report = {
            'checked_count': len(checked_ids),
            'missing_cars': missing_cars,
            'extra_vins': extra_not_in_system,
            'discrepancy_count': len(missing_cars) + len(extra_not_in_system),
        }

    return render(request, 'sales/inventory_reconciliation.html', {
        'cars': cars,
        'checked_ids': checked_ids,
        'extra_vins_text': extra_vins_text,
        'report': report,
    })


@require_interface_access('can_access_reports')
def vouchers_reports(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    return render(request, 'sales/reports_vouchers.html')


def _normalize_bank_header(value):
    text = (str(value or '')).strip().lower()
    text = text.replace('\u200f', '').replace('\u200e', '')
    text = text.replace('-', '_').replace(' ', '_')
    while '__' in text:
        text = text.replace('__', '_')
    return text


def _parse_statement_decimal(value):
    if value is None or value == '':
        return None

    if isinstance(value, Decimal):
        return value

    if isinstance(value, (int, float)):
        return Decimal(str(value))

    normalized = str(value).strip().replace(',', '')
    if not normalized:
        return None

    try:
        return Decimal(normalized)
    except Exception:
        return None


def _parse_statement_date(value):
    if value is None or value == '':
        return None

    if isinstance(value, datetime):
        return value.date()

    if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day'):
        return value

    text = str(value).strip()
    if not text:
        return None

    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    return None


def _parse_bank_statement_rows(file_obj):
    workbook = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    worksheet = workbook.active

    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        workbook.close()
        return [], ['ملف كشف الحساب فارغ.']

    aliases = {
        'voucher_number': {
            'voucher', 'voucher_number', 'voucher_no', 'reference', 'reference_no', 'ref', 'receipt_number',
            'رقم_السند', 'رقم_المرجع', 'مرجع', 'السند',
        },
        'transaction_date': {
            'date', 'transaction_date', 'voucher_date', 'booking_date', 'value_date',
            'التاريخ', 'تاريخ_الحركة', 'تاريخ_القيد',
        },
        'amount': {
            'amount', 'value', 'net_amount', 'transaction_amount',
            'المبلغ', 'القيمة', 'صافي_المبلغ',
        },
        'currency': {
            'currency', 'curr', 'عملة', 'العملة',
        },
        'description': {
            'description', 'memo', 'note', 'notes', 'details',
            'البيان', 'الوصف', 'ملاحظات',
        },
    }

    indexes = {}
    normalized_headers = [_normalize_bank_header(col) for col in header_row]
    for idx, header in enumerate(normalized_headers):
        for field_name, field_aliases in aliases.items():
            if header in field_aliases and field_name not in indexes:
                indexes[field_name] = idx

    if 'amount' not in indexes:
        workbook.close()
        return [], ['تعذر تحديد عمود المبلغ داخل ملف كشف الحساب.']

    parsed_rows = []
    warnings = []
    for line_no, row in enumerate(rows_iter, start=2):
        if not row or all(cell in (None, '') for cell in row):
            continue

        def _cell(field_name, default=''):
            col_idx = indexes.get(field_name)
            if col_idx is None or col_idx >= len(row):
                return default
            return row[col_idx]

        amount = _parse_statement_decimal(_cell('amount'))
        if amount is None:
            warnings.append(f'تجاهل السطر {line_no}: قيمة المبلغ غير صالحة.')
            continue

        raw_currency = str(_cell('currency', 'SR') or 'SR').strip().upper()
        currency = raw_currency if raw_currency in {'SR', '$', '£'} else 'SR'

        parsed_rows.append({
            'line_no': line_no,
            'voucher_number': str(_cell('voucher_number', '') or '').strip().upper(),
            'transaction_date': _parse_statement_date(_cell('transaction_date')),
            'amount': amount,
            'currency': currency,
            'description': str(_cell('description', '') or '').strip(),
        })

    workbook.close()
    return parsed_rows, warnings


def _build_bank_statement_template_response():
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Bank Statement Template'
    worksheet.append(['رقم السند', 'التاريخ', 'المبلغ', 'العملة', 'البيان'])
    worksheet.append(['SDF-20260307-1001', timezone.localdate().isoformat(), '25000', 'SR', 'حوالة بيع سيارة'])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="bank_statement_template.xlsx"'
    workbook.save(response)
    return response


@require_interface_access('can_access_reports')
def bank_reconciliation(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    if request.method == 'GET' and request.GET.get('download_template') == '1':
        return _build_bank_statement_template_response()

    form = BankReconciliationUploadForm()
    result = None

    if request.method == 'POST':
        form = BankReconciliationUploadForm(request.POST, request.FILES)
        if form.is_valid():
            statement_rows, parse_warnings = _parse_bank_statement_rows(form.cleaned_data['statement_file'])
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')

            vouchers_qs = FinanceVoucher.objects.all()
            if start_date:
                vouchers_qs = vouchers_qs.filter(voucher_date__gte=start_date)
            if end_date:
                vouchers_qs = vouchers_qs.filter(voucher_date__lte=end_date)

            system_vouchers = list(vouchers_qs.order_by('voucher_date', 'id'))
            voucher_number_map = defaultdict(list)
            for voucher in system_vouchers:
                voucher_number_map[(voucher.voucher_number or '').strip().upper()].append(voucher)

            matched_rows = []
            statement_only_rows = []
            matched_voucher_ids = set()

            for row in statement_rows:
                match = None
                statement_voucher_number = row['voucher_number']

                if statement_voucher_number:
                    for candidate in voucher_number_map.get(statement_voucher_number, []):
                        if candidate.id not in matched_voucher_ids:
                            match = candidate
                            break

                if match is None and row['transaction_date'] is not None:
                    target_amount_sr = _to_sr(row['amount'], row['currency'])
                    for candidate in system_vouchers:
                        if candidate.id in matched_voucher_ids:
                            continue
                        if candidate.voucher_date != row['transaction_date']:
                            continue
                        if _to_sr(candidate.amount, candidate.currency) == target_amount_sr:
                            match = candidate
                            break

                if match is not None:
                    matched_voucher_ids.add(match.id)
                    matched_rows.append({
                        'statement': row,
                        'voucher': match,
                        'amount_sr': _to_sr(match.amount, match.currency),
                    })
                else:
                    row['amount_sr'] = _to_sr(row['amount'], row['currency'])
                    statement_only_rows.append(row)

            system_only_rows = [voucher for voucher in system_vouchers if voucher.id not in matched_voucher_ids]

            result = {
                'warnings': parse_warnings,
                'statement_rows_count': len(statement_rows),
                'system_rows_count': len(system_vouchers),
                'matched_count': len(matched_rows),
                'statement_only_count': len(statement_only_rows),
                'system_only_count': len(system_only_rows),
                'matched_rows': matched_rows,
                'statement_only_rows': statement_only_rows,
                'system_only_rows': system_only_rows,
                'matched_total_sr': sum((item['amount_sr'] for item in matched_rows), Decimal('0')),
                'statement_only_total_sr': sum((item['amount_sr'] for item in statement_only_rows), Decimal('0')),
                'system_only_total_sr': sum((_to_sr(v.amount, v.currency) for v in system_only_rows), Decimal('0')),
                'start_date': start_date,
                'end_date': end_date,
            }

    return render(request, 'sales/bank_reconciliation.html', {
        'form': form,
        'result': result,
    })


@require_interface_access('can_access_reports')
def financial_consistency_checker(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    tenant_alias = (get_current_tenant_db_alias() or '').strip()
    if not tenant_alias.startswith('tenant_'):
        raise PermissionDenied('لا توجد بيئة معرض مفعلة لفحص التوازن المالي.')

    report = build_financial_consistency_report(alias=tenant_alias)
    return render(request, 'sales/financial_consistency_checker.html', {
        'report': report,
        'tenant_alias': tenant_alias,
    })


@require_interface_access('can_access_reports')
def vouchers_list(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    tenant_alias = _require_tenant_alias_or_raise('لا توجد بيئة معرض مفعلة لعرض السندات.')
    ensure_default_chart_of_accounts(alias=tenant_alias)
    ensure_default_financial_containers(alias=tenant_alias)

    for payment in DebtPayment.objects.select_related('sale__car', 'sale__customer'):
        settlement_container = get_default_financial_container(
            alias=tenant_alias,
            preferred_type=FinancialContainer.TYPE_MAIN_CASH,
            currency=payment.sale.car.currency,
        )
        settlement_credit_choice = FinanceVoucher.ACCOUNT_CASH_BOX
        if settlement_container and settlement_container.container_type == FinancialContainer.TYPE_BANK:
            settlement_credit_choice = FinanceVoucher.ACCOUNT_BANK

        FinanceVoucher.objects.get_or_create(
            voucher_number=payment.receipt_number,
            defaults={
                'voucher_type': 'settlement',
                'voucher_date': payment.payment_date,
                'person_name': payment.sale.customer.name,
                'amount': payment.paid_amount,
                'currency': payment.sale.car.currency,
                'reason': f"تسديد مديونية سيارة {payment.sale.car.brand} {payment.sale.car.model_name}",
                'linked_car': payment.sale.car,
                'financial_container': settlement_container,
                'debit_account': FinanceVoucher.ACCOUNT_CASH_BOX,
                'credit_account': settlement_credit_choice,
            }
        )

    selected_type = request.GET.get('type', 'all')
    vouchers = FinanceVoucher.objects.select_related(
        'maintenance_record',
        'maintenance_record__car',
        'financial_container',
        'linked_car',
    ).all()
    if selected_type != 'all':
        vouchers = vouchers.filter(voucher_type=selected_type)

    type_label_map = {
        'all': 'الكل',
        'payment': 'سند دفع',
        'receipt': 'سند قبض',
        'operating': 'مصروفات تشغيلية',
        'settlement': 'سند تسديد',
        'maintenance': 'قيد صيانة سيارات',
    }

    context = {
        'vouchers': vouchers.order_by('-voucher_date', '-created_at'),
        'selected_type': selected_type,
        'selected_type_label': type_label_map.get(selected_type, 'الكل'),
    }
    return render(request, 'sales/vouchers_list.html', context)


@require_interface_access('can_access_reports')
def receipt_voucher(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    tenant_alias = _require_tenant_alias_or_raise('لا توجد بيئة معرض مفعلة لسند القبض.')
    ensure_default_chart_of_accounts(alias=tenant_alias)
    ensure_default_financial_containers(alias=tenant_alias)

    now = timezone.now()
    default_number = f"QBD-{now.strftime('%Y%m%d-%H%M')}"
    default_date = timezone.localdate()
    if request.method == 'POST':
        form = ReceiptVoucherForm(request.POST, request.FILES, tenant_alias=tenant_alias)
        if form.is_valid():
            container = form.cleaned_data['financial_container']
            currency = form.cleaned_data['receipt_currency']
            if container.currency != currency:
                form.add_error('receipt_currency', 'عملة السند يجب أن تطابق عملة الوعاء المالي.')
            else:
                debit_choice = (
                    FinanceVoucher.ACCOUNT_BANK
                    if container.container_type == FinancialContainer.TYPE_BANK
                    else FinanceVoucher.ACCOUNT_CASH_BOX
                )
                FinanceVoucher.objects.create(
                    voucher_type='receipt',
                    voucher_number=form.cleaned_data['voucher_number'],
                    voucher_date=form.cleaned_data['voucher_date'],
                    person_name=form.cleaned_data['receiver_name'],
                    amount=form.cleaned_data['receipt_amount'],
                    currency=currency,
                    reason=form.cleaned_data['receipt_reason'],
                    linked_car=form.cleaned_data.get('linked_car'),
                    financial_container=container,
                    supporting_document=form.cleaned_data.get('supporting_document'),
                    debit_account=debit_choice,
                    credit_account=FinanceVoucher.ACCOUNT_NONE,
                )
                return redirect(f"{reverse('receipt_voucher')}?saved=1")
    else:
        form = ReceiptVoucherForm(initial={
            'voucher_number': default_number,
            'voucher_date': default_date,
        }, tenant_alias=tenant_alias)

    return render(request, 'sales/receipt_voucher.html', {
        'form': form,
        'receipt_ok_url': reverse('receipt_voucher'),
    })


@require_interface_access('can_access_reports')
def payment_voucher(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    tenant_alias = _require_tenant_alias_or_raise('لا توجد بيئة معرض مفعلة لسند الدفع.')
    ensure_default_chart_of_accounts(alias=tenant_alias)
    ensure_default_financial_containers(alias=tenant_alias)

    now = timezone.now()
    default_number = f"SDF-{now.strftime('%Y%m%d-%H%M')}"
    default_date = timezone.localdate()
    if request.method == 'POST':
        form = PaymentVoucherForm(request.POST, request.FILES, tenant_alias=tenant_alias)
        if form.is_valid():
            container = form.cleaned_data['financial_container']
            currency = form.cleaned_data['payment_currency']
            if container.currency != currency:
                form.add_error('payment_currency', 'عملة السند يجب أن تطابق عملة الوعاء المالي.')
            else:
                credit_choice = (
                    FinanceVoucher.ACCOUNT_BANK
                    if container.container_type == FinancialContainer.TYPE_BANK
                    else FinanceVoucher.ACCOUNT_CASH_BOX
                )
                FinanceVoucher.objects.create(
                    voucher_type='payment',
                    voucher_number=form.cleaned_data['voucher_number'],
                    voucher_date=form.cleaned_data['voucher_date'],
                    person_name=form.cleaned_data['payer_name'],
                    amount=form.cleaned_data['payment_amount'],
                    currency=currency,
                    reason=form.cleaned_data['payment_reason'],
                    linked_car=form.cleaned_data.get('linked_car'),
                    financial_container=container,
                    supporting_document=form.cleaned_data.get('supporting_document'),
                    debit_account=FinanceVoucher.ACCOUNT_NONE,
                    credit_account=credit_choice,
                )
                return redirect(f"{reverse('payment_voucher')}?saved=1")
    else:
        form = PaymentVoucherForm(initial={
            'voucher_number': default_number,
            'voucher_date': default_date,
        }, tenant_alias=tenant_alias)

    return render(request, 'sales/payment_voucher.html', {
        'form': form,
        'payment_ok_url': reverse('payment_voucher'),
    })


@require_interface_access('can_access_reports')
def operating_expenses_voucher(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    tenant_alias = _require_tenant_alias_or_raise('لا توجد بيئة معرض مفعلة لسند المصروف التشغيلي.')
    ensure_default_chart_of_accounts(alias=tenant_alias)
    ensure_default_financial_containers(alias=tenant_alias)

    now = timezone.now()
    default_number = f"MST-{now.strftime('%Y%m%d-%H%M')}"
    default_date = timezone.localdate()
    if request.method == 'POST':
        form = OperatingExpenseVoucherForm(request.POST, request.FILES, tenant_alias=tenant_alias)
        if form.is_valid():
            container = form.cleaned_data['financial_container']
            currency = form.cleaned_data['expense_currency']
            if container.currency != currency:
                form.add_error('expense_currency', 'عملة السند يجب أن تطابق عملة الوعاء المالي.')
            else:
                credit_choice = (
                    FinanceVoucher.ACCOUNT_BANK
                    if container.container_type == FinancialContainer.TYPE_BANK
                    else FinanceVoucher.ACCOUNT_CASH_BOX
                )
                FinanceVoucher.objects.create(
                    voucher_type='operating',
                    voucher_number=form.cleaned_data['voucher_number'],
                    voucher_date=form.cleaned_data['voucher_date'],
                    person_name=form.cleaned_data['expense_name'],
                    amount=form.cleaned_data['expense_amount'],
                    currency=currency,
                    reason=form.cleaned_data['allowance_reason'],
                    linked_car=form.cleaned_data.get('linked_car'),
                    financial_container=container,
                    supporting_document=form.cleaned_data.get('supporting_document'),
                    debit_account=FinanceVoucher.ACCOUNT_OPERATING_EXPENSES,
                    credit_account=credit_choice,
                )
                return redirect(f"{reverse('operating_expenses_voucher')}?saved=1")
    else:
        form = OperatingExpenseVoucherForm(initial={
            'voucher_number': default_number,
            'voucher_date': default_date,
        }, tenant_alias=tenant_alias)

    return render(request, 'sales/operating_expenses_voucher.html', {
        'form': form,
        'operating_ok_url': reverse('operating_expenses_voucher'),
    })


@require_interface_access('can_access_reports')
def cash_flow_projection(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    today = timezone.localdate()
    try:
        horizon_days = int(request.GET.get('horizon_days', 56))
    except (TypeError, ValueError):
        horizon_days = 56

    if horizon_days not in {14, 30, 56, 90}:
        horizon_days = 56

    horizon_date = today + timedelta(days=horizon_days)

    pending_sales = (
        Sale.objects.select_related('car', 'customer')
        .filter(amount_paid__lt=F('sale_price'))
        .order_by('debt_due_date', 'sale_date')
    )

    rows = []
    weekly_projection_map = defaultdict(Decimal)
    monthly_projection_map = defaultdict(Decimal)

    total_outstanding_sr = Decimal('0')
    overdue_total_sr = Decimal('0')
    projected_total_sr = Decimal('0')
    projected_within_7_sr = Decimal('0')
    projected_within_30_sr = Decimal('0')

    for sale in pending_sales:
        sale_day = timezone.localtime(sale.sale_date).date() if sale.sale_date else today
        due_date = sale.debt_due_date or (sale_day + timedelta(days=30))
        remaining_amount = sale.remaining_amount
        remaining_amount_sr = _to_sr(remaining_amount, getattr(sale.car, 'currency', 'SR'))

        total_outstanding_sr += remaining_amount_sr

        is_overdue = due_date < today
        if is_overdue:
            overdue_total_sr += remaining_amount_sr

        projected_date = today if is_overdue else due_date
        in_horizon = projected_date <= horizon_date

        if in_horizon:
            projected_total_sr += remaining_amount_sr
            if projected_date <= today + timedelta(days=7):
                projected_within_7_sr += remaining_amount_sr
            if projected_date <= today + timedelta(days=30):
                projected_within_30_sr += remaining_amount_sr

            week_start = projected_date - timedelta(days=projected_date.weekday())
            weekly_projection_map[week_start] += remaining_amount_sr
            monthly_projection_map[projected_date.strftime('%Y-%m')] += remaining_amount_sr

        rows.append({
            'sale': sale,
            'due_date': due_date,
            'projected_date': projected_date,
            'is_overdue': is_overdue,
            'in_horizon': in_horizon,
            'remaining_amount': remaining_amount,
            'remaining_amount_sr': remaining_amount_sr,
        })

    rows.sort(key=lambda item: (item['projected_date'], item['sale'].id))

    weekly_projection = [
        {'week_start': week_start, 'amount_sr': amount}
        for week_start, amount in sorted(weekly_projection_map.items(), key=lambda item: item[0])
    ]
    monthly_projection = [
        {'month_label': month_label, 'amount_sr': amount}
        for month_label, amount in sorted(monthly_projection_map.items(), key=lambda item: item[0])
    ]

    return render(request, 'sales/cash_flow_projection.html', {
        'today': today,
        'horizon_days': horizon_days,
        'horizon_date': horizon_date,
        'rows': rows,
        'weekly_projection': weekly_projection,
        'monthly_projection': monthly_projection,
        'total_outstanding_sr': total_outstanding_sr,
        'overdue_total_sr': overdue_total_sr,
        'projected_total_sr': projected_total_sr,
        'projected_within_7_sr': projected_within_7_sr,
        'projected_within_30_sr': projected_within_30_sr,
    })


REPORT_CURRENCY = 'SR'
CURRENCY_RATES_TO_SR = {
    'SR': Decimal('1'),
    '$': Decimal('3.75'),
    '£': Decimal('4.80'),
}


def _to_sr(amount, currency='SR'):
    normalized_amount = amount or Decimal('0')
    rate = CURRENCY_RATES_TO_SR.get(currency, Decimal('1'))
    return Decimal(normalized_amount) * rate


def _aggregate_sales_time_series(sales, granularity='month'):
    buckets = defaultdict(Decimal)
    for sale in sales:
        if granularity == 'day':
            label = sale.sale_date.date().strftime('%Y-%m-%d')
        else:
            label = sale.sale_date.strftime('%Y-%m')
        buckets[label] += _to_sr(sale.sale_price, getattr(sale.car, 'currency', 'SR'))

    labels = sorted(buckets.keys())
    values = [float(buckets[label]) for label in labels]
    return labels, values


def _month_key(dt):
    return dt.strftime('%Y-%m')


def _last_month_labels(total_months=6):
    current = timezone.localdate().replace(day=1)
    labels = []
    for _ in range(total_months):
        labels.append(current.strftime('%Y-%m'))
        if current.month == 1:
            current = current.replace(year=current.year - 1, month=12, day=1)
        else:
            current = current.replace(month=current.month - 1, day=1)
    return list(reversed(labels))


def _build_monthly_comparison_series():
    labels = _last_month_labels(6)
    values_map = {label: Decimal('0') for label in labels}

    sales = Sale.objects.select_related('car').all()
    for sale in sales:
        key = _month_key(sale.sale_date)
        if key in values_map:
            values_map[key] += _to_sr(sale.sale_price, getattr(sale.car, 'currency', 'SR'))

    values = [float(values_map[label]) for label in labels]
    return labels, values


def _build_maintenance_monthly_series(total_months=6):
    labels = _last_month_labels(total_months)
    values_map = {label: Decimal('0') for label in labels}

    maintenance_records = CarMaintenance.objects.select_related('car').all()
    for maintenance in maintenance_records:
        key = maintenance.operation_date.strftime('%Y-%m')
        if key in values_map:
            values_map[key] += _to_sr(maintenance.amount, getattr(maintenance.car, 'currency', 'SR'))

    series = [{'label': label, 'value': values_map[label]} for label in labels]
    current_month_total = series[-1]['value'] if series else Decimal('0')
    previous_values = [item['value'] for item in series[:-1]]
    previous_average = (
        sum(previous_values, Decimal('0')) / Decimal(len(previous_values))
        if previous_values
        else Decimal('0')
    )
    return series, current_month_total, previous_average


def _build_financial_report_context(query_params):
    period = query_params.get('period', 'custom')
    reference_date_str = query_params.get('reference_date')
    start_date = query_params.get('start_date')
    end_date = query_params.get('end_date')

    today = timezone.localdate()
    reference_date = None

    if reference_date_str:
        try:
            reference_date = datetime.strptime(reference_date_str, '%Y-%m-%d').date()
        except ValueError:
            reference_date = None

    if period in {'daily', 'monthly', 'yearly'}:
        base_date = reference_date or today
        if period == 'daily':
            start_date = base_date.isoformat()
            end_date = base_date.isoformat()
        elif period == 'monthly':
            month_start = base_date.replace(day=1)
            if month_start.month == 12:
                next_month_start = month_start.replace(year=month_start.year + 1, month=1, day=1)
            else:
                next_month_start = month_start.replace(month=month_start.month + 1, day=1)
            month_end = next_month_start - timedelta(days=1)
            start_date = month_start.isoformat()
            end_date = month_end.isoformat()
        else:
            start_date = base_date.replace(month=1, day=1).isoformat()
            end_date = base_date.replace(month=12, day=31).isoformat()
    elif period == 'custom':
        if not start_date and not end_date:
            month_start = today.replace(day=1)
            start_date = month_start.isoformat()
            end_date = today.isoformat()
    else:
        period = 'custom'

    sales = Sale.objects.select_related('car', 'customer').all()
    expenses = Expense.objects.all()
    general_expenses = GeneralExpense.objects.all()
    maintenance_records = CarMaintenance.objects.select_related('car').all()
    debt_payments = DebtPayment.objects.select_related('sale').all()
    vouchers = FinanceVoucher.objects.all()

    if start_date:
        sales = sales.filter(sale_date__date__gte=start_date)
        expenses = expenses.filter(date__gte=start_date)
        general_expenses = general_expenses.filter(expense_date__gte=start_date)
        maintenance_records = maintenance_records.filter(operation_date__gte=start_date)
        debt_payments = debt_payments.filter(payment_date__gte=start_date)
        vouchers = vouchers.filter(voucher_date__gte=start_date)
    if end_date:
        sales = sales.filter(sale_date__date__lte=end_date)
        expenses = expenses.filter(date__lte=end_date)
        general_expenses = general_expenses.filter(expense_date__lte=end_date)
        maintenance_records = maintenance_records.filter(operation_date__lte=end_date)
        debt_payments = debt_payments.filter(payment_date__lte=end_date)
        vouchers = vouchers.filter(voucher_date__lte=end_date)

    sales_list = list(sales)
    expense_list = list(expenses)
    general_expense_list = list(general_expenses)
    maintenance_list = list(maintenance_records)
    debt_payment_list = list(debt_payments)
    voucher_list = list(vouchers)

    sold_car_ids = [sale.car_id for sale in sales_list]
    sold_maintenance_totals = {
        row['car_id']: row['total'] or Decimal('0')
        for row in CarMaintenance.objects.filter(car_id__in=sold_car_ids)
        .values('car_id')
        .annotate(total=Sum('amount'))
    } if sold_car_ids else {}

    total_sales = Decimal('0')
    total_cost_sold = Decimal('0')
    total_additional_cost_sold = Decimal('0')
    cash_from_sales = Decimal('0')
    outstanding_debts = Decimal('0')

    for sale in sales_list:
        currency = getattr(sale.car, 'currency', 'SR')
        sale_price_sr = _to_sr(sale.sale_price, currency)
        maintenance_amount = sold_maintenance_totals.get(sale.car_id, Decimal('0'))
        additional_cost_amount = (
            (sale.car.customs_cost or Decimal('0'))
            + (sale.car.transport_cost or Decimal('0'))
            + (sale.car.commission_cost or Decimal('0'))
        )
        full_cost_price = (sale.car.cost_price or Decimal('0')) + additional_cost_amount + maintenance_amount
        cost_price_sr = _to_sr(full_cost_price, currency)
        amount_paid_sr = _to_sr(sale.amount_paid, currency)
        remaining_sr = sale_price_sr - amount_paid_sr

        total_sales += sale_price_sr
        total_cost_sold += cost_price_sr
        total_additional_cost_sold += _to_sr(additional_cost_amount, currency)
        cash_from_sales += amount_paid_sr
        outstanding_debts += remaining_sr

        sale.sale_price_sr = sale_price_sr
        sale.cost_price_sr = cost_price_sr
        sale.base_cost_sr = _to_sr(sale.car.cost_price or Decimal('0'), currency)
        sale.customs_cost_sr = _to_sr(sale.car.customs_cost or Decimal('0'), currency)
        sale.transport_cost_sr = _to_sr(sale.car.transport_cost or Decimal('0'), currency)
        sale.commission_cost_sr = _to_sr(sale.car.commission_cost or Decimal('0'), currency)
        sale.additional_cost_sr = _to_sr(additional_cost_amount, currency)
        sale.maintenance_total_sr = _to_sr(maintenance_amount, currency)
        sale.profit_sr = sale_price_sr - cost_price_sr
        sale.accrual_profit_sr = sale.profit_sr
        sale.amount_paid_sr = amount_paid_sr
        sale.uncollected_sr = remaining_sr
        sale.cash_realized_profit_sr = amount_paid_sr - cost_price_sr

    gross_profit = total_sales - total_cost_sold

    operating_expenses_total = sum((expense.amount for expense in expense_list), Decimal('0'))
    general_expenses_total = sum(
        (_to_sr(item.amount, item.currency) for item in general_expense_list),
        Decimal('0'),
    )
    maintenance_total = sum(
        (_to_sr(item.amount, item.car.currency) for item in maintenance_list),
        Decimal('0'),
    )
    voucher_operating_total = sum(
        (_to_sr(v.amount, v.currency) for v in voucher_list if v.voucher_type == 'operating'),
        Decimal('0'),
    )
    total_expenses = operating_expenses_total + general_expenses_total + maintenance_total + voucher_operating_total
    final_net_profit = gross_profit - total_expenses

    available_cars = list(Car.objects.filter(is_sold=False).only('id', 'cost_price', 'currency'))
    available_car_ids = [car.id for car in available_cars]
    available_maintenance_totals = {
        row['car_id']: row['total'] or Decimal('0')
        for row in CarMaintenance.objects.filter(car_id__in=available_car_ids)
        .values('car_id')
        .annotate(total=Sum('amount'))
    } if available_car_ids else {}

    inventory_value = sum(
        (
            _to_sr(
                (
                    (car.cost_price or Decimal('0'))
                    + (car.customs_cost or Decimal('0'))
                    + (car.transport_cost or Decimal('0'))
                    + (car.commission_cost or Decimal('0'))
                    + available_maintenance_totals.get(car.id, Decimal('0'))
                ),
                car.currency,
            )
            for car in available_cars
        ),
        Decimal('0'),
    )
    available_count = len(available_cars)
    sold_count = len(sales_list)

    cash_manual_receipts = sum(
        (_to_sr(v.amount, v.currency) for v in voucher_list if v.voucher_type == 'receipt'),
        Decimal('0'),
    )
    cash_settlements = sum(
        (_to_sr(v.amount, v.currency) for v in voucher_list if v.voucher_type == 'settlement'),
        Decimal('0'),
    )
    total_cash_in = cash_from_sales + cash_manual_receipts + cash_settlements

    voucher_payments = sum(
        (_to_sr(v.amount, v.currency) for v in voucher_list if v.voucher_type == 'payment'),
        Decimal('0'),
    )
    total_cash_out = total_expenses + voucher_payments
    net_cash_flow = total_cash_in - total_cash_out
    accrual_basis_profit = final_net_profit
    cash_basis_profit = net_cash_flow
    accrual_cash_gap = accrual_basis_profit - cash_basis_profit
    receivables_impact_ratio = Decimal('0')
    if total_sales > 0:
        receivables_impact_ratio = (outstanding_debts / total_sales) * Decimal('100')

    debts_collected = sum(
        (_to_sr(payment.paid_amount, payment.sale.car.currency) for payment in debt_payment_list),
        Decimal('0'),
    )

    chart_granularity = query_params.get('chart_group')
    if chart_granularity not in {'day', 'month'}:
        chart_granularity = 'day' if period == 'daily' else 'month'

    line_labels, line_values = _aggregate_sales_time_series(sales_list, chart_granularity)
    bar_labels, bar_values = _build_monthly_comparison_series()
    maintenance_monthly_series, maintenance_current_month, maintenance_previous_avg = _build_maintenance_monthly_series(6)

    doughnut_profit = float(max(final_net_profit, Decimal('0')))
    doughnut_expenses = float(total_expenses)
    doughnut_debts = float(max(outstanding_debts, Decimal('0')))

    context = {
        'gross_profit': gross_profit,
        'net_profit': final_net_profit,
        'total_sales': total_sales,
        'total_expenses': total_expenses,
        'final_net_profit': final_net_profit,
        'total_additional_cost_sold': total_additional_cost_sold,
        'operating_expenses_total': operating_expenses_total,
        'general_expenses_total': general_expenses_total,
        'maintenance_total': maintenance_total,
        'voucher_operating_total': voucher_operating_total,
        'start_date': start_date,
        'end_date': end_date,
        'period': period,
        'reference_date': reference_date_str or '',
        'sold_count': sold_count,
        'available_count': available_count,
        'sales_list': sales_list,
        'inventory_value': inventory_value,
        'cash_from_sales': cash_from_sales,
        'cash_manual_receipts': cash_manual_receipts,
        'cash_settlements': cash_settlements,
        'total_cash_in': total_cash_in,
        'voucher_payments': voucher_payments,
        'total_cash_out': total_cash_out,
        'net_cash_flow': net_cash_flow,
        'accrual_basis_profit': accrual_basis_profit,
        'cash_basis_profit': cash_basis_profit,
        'accrual_cash_gap': accrual_cash_gap,
        'receivables_impact_ratio': receivables_impact_ratio,
        'outstanding_debts': outstanding_debts,
        'debts_collected': debts_collected,
        'report_currency': REPORT_CURRENCY,
        'chart_granularity': chart_granularity,
        'line_chart': {
            'labels': line_labels,
            'values': line_values,
        },
        'doughnut_chart': {
            'labels': ['أرباح', 'مصاريف', 'ديون'],
            'values': [doughnut_profit, doughnut_expenses, doughnut_debts],
        },
        'bar_chart': {
            'labels': bar_labels,
            'values': bar_values,
        },
        'maintenance_monthly_series': maintenance_monthly_series,
        'maintenance_current_month': maintenance_current_month,
        'maintenance_previous_avg': maintenance_previous_avg,
    }
    return context


@require_interface_access('can_access_reports')
def financial_reports_detail(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    context = _build_financial_report_context(request.GET)
    return render(request, 'sales/reports_financial.html', context)


@require_interface_access('can_access_reports')
def export_financial_excel(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    context = _build_financial_report_context(request.GET)

    reports_dir = Path(settings.MEDIA_ROOT) / 'reports'
    reports_dir.mkdir(parents=True, exist_ok=True)

    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    file_name = f'financial_report_{timestamp}.xlsx'
    file_path = reports_dir / file_name

    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = 'الملخص المالي'

    summary_sheet.append(['البند', 'القيمة'])
    summary_sheet.append(['الفترة', context.get('period', '')])
    summary_sheet.append(['من تاريخ', context.get('start_date', '')])
    summary_sheet.append(['إلى تاريخ', context.get('end_date', '')])
    summary_sheet.append(['إجمالي المبيعات', float(context['total_sales'])])
    summary_sheet.append(['الربح الأولي', float(context['gross_profit'])])
    summary_sheet.append(['إجمالي المصروفات', float(context['total_expenses'])])
    summary_sheet.append(['المصروفات العامة', float(context['general_expenses_total'])])
    summary_sheet.append(['مصروفات الصيانة', float(context['maintenance_total'])])
    summary_sheet.append(['صافي الربح النهائي', float(context['final_net_profit'])])
    summary_sheet.append(['الربح على أساس الاستحقاق', float(context['accrual_basis_profit'])])
    summary_sheet.append(['الربح على الأساس النقدي', float(context['cash_basis_profit'])])
    summary_sheet.append(['فجوة الاستحقاق مقابل النقدي', float(context['accrual_cash_gap'])])
    summary_sheet.append(['قيمة المخزون الحالي', float(context['inventory_value'])])
    summary_sheet.append(['عدد السيارات المباعة', context['sold_count']])
    summary_sheet.append(['عدد السيارات المتاحة', context['available_count']])
    summary_sheet.append(['إجمالي الكاش الداخل', float(context['total_cash_in'])])
    summary_sheet.append(['إجمالي الكاش الخارج', float(context['total_cash_out'])])
    summary_sheet.append(['صافي التدفق النقدي', float(context['net_cash_flow'])])
    summary_sheet.append(['الديون المتبقية', float(context['outstanding_debts'])])
    summary_sheet.append(['نسبة تأثير الذمم المدينة %', float(context['receivables_impact_ratio'])])
    summary_sheet.append(['التسديدات المحصلة', float(context['debts_collected'])])

    sales_sheet = workbook.create_sheet(title='تفاصيل المبيعات')
    sales_sheet.append([
        'السيارة',
        'العميل',
        'التكلفة الأساسية',
        'التكاليف الإضافية',
        'الصيانة',
        'إجمالي التكلفة',
        'سعر البيع',
        'الربح (استحقاق)',
        'المحصل نقدًا',
        'الربح النقدي',
        'تاريخ البيع',
    ])
    for sale in context['sales_list']:
        sales_sheet.append([
            f'{sale.car.brand} ({sale.car.vin})',
            sale.customer.name,
            float(sale.base_cost_sr),
            float(sale.additional_cost_sr),
            float(sale.maintenance_total_sr),
            float(sale.cost_price_sr),
            float(sale.sale_price_sr),
            float(sale.profit_sr),
            float(sale.amount_paid_sr),
            float(sale.cash_realized_profit_sr),
            sale.sale_date.strftime('%Y-%m-%d %H:%M'),
        ])

    workbook.save(file_path)
    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=file_name)


@require_interface_access('can_access_reports')
def financial_report_charts_data(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    context = _build_financial_report_context(request.GET)
    return JsonResponse({
        'report_currency': context['report_currency'],
        'chart_granularity': context['chart_granularity'],
        'line_chart': context['line_chart'],
        'doughnut_chart': context['doughnut_chart'],
        'bar_chart': context['bar_chart'],
        'updated_at': timezone.now().isoformat(),
    })


def _normalize_operation_text(raw_text):
    text = (raw_text or '').strip()
    text = re.sub(r'\s*\(سابق\)', '', text)
    text = re.sub(r'\s*\(ID:\s*\d+\)', '', text)
    text = re.sub(r'\s*ID:\s*\d+', '', text)

    if text.startswith('إضافة:'):
        text = text.replace('إضافة:', 'إضافة', 1).strip()
    elif text.startswith('تعديل:'):
        text = text.replace('تعديل:', 'تعديل', 1).strip()
    elif text.startswith('حذف:'):
        text = text.replace('حذف:', 'حذف', 1).strip()

    return text


@require_interface_access('can_access_timeline')
def timeline_view(request):
    logs = OperationLog.objects.select_related('user').all()

    for log in logs:
        log.display_operation = _normalize_operation_text(log.operation)

    return render(request, 'sales/timeline.html', {'logs': logs})


@require_interface_access('can_access_timeline')
def export_timeline_excel(request):
    logs = OperationLog.objects.select_related('user').all()

    reports_dir = Path(settings.MEDIA_ROOT) / 'reports'
    reports_dir.mkdir(parents=True, exist_ok=True)

    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    file_name = f'timeline_report_{timestamp}.xlsx'
    file_path = reports_dir / file_name

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'الجدول الزمني'
    sheet.append(['العملية', 'الحساب', 'التاريخ والوقت'])

    for log in logs:
        sheet.append([
            _normalize_operation_text(log.operation),
            log.user.username if log.user else 'غير معروف',
            timezone.localtime(log.created_at).strftime('%Y-%m-%d %H:%M:%S'),
        ])

    workbook.save(file_path)
    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=file_name)


@ensure_csrf_cookie
def user_login(request):
    next_url = request.GET.get('next') or request.POST.get('next') or ''

    if request.user.is_authenticated:
        if next_url.startswith('/admin/') and not (
            request.user.is_staff or
            request.user.is_superuser or
            _is_platform_owner_session(request)
        ):
            messages.error(request, 'حسابك الحالي لا يملك صلاحية دخول لوحة الإدارة.')
            return redirect('home')

        if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()):
            return redirect(next_url)
        return redirect('home')

    if request.method == 'POST':
        form = TenantLoginForm(request.POST)
        if form.is_valid():
            tenant_id_raw = (form.cleaned_data.get('tenant_id') or '').strip()
            tenant_key = (form.cleaned_data.get('tenant_key') or '').strip()
            username = (form.cleaned_data.get('username') or '').strip()
            password = (form.cleaned_data.get('password') or '').strip()

            # Hidden platform-admin path: when tenant credentials are omitted,
            # allow login only for default-db superuser/staff accounts.
            if not tenant_id_raw and not tenant_key:
                client_ip = _extract_client_ip(request)
                hidden_login_key = _hidden_platform_login_cache_key(username, client_ip)
                current_attempts = int(cache.get(hidden_login_key) or 0)

                if current_attempts >= HIDDEN_PLATFORM_LOGIN_MAX_ATTEMPTS:
                    _write_hidden_platform_login_failure_audit(
                        request,
                        username,
                        reason='rate_limited_hidden_platform_login',
                    )
                    form.add_error(
                        None,
                        'تم تقييد محاولات الدخول مؤقتًا. الرجاء المحاولة لاحقًا.',
                    )
                    return render(
                        request,
                        'admin/login.html',
                        {
                            'form': form,
                            'next': next_url,
                            'google_oauth_enabled': _google_oauth_enabled(),
                        },
                    )

                owner = User.objects.using('default').filter(
                    username=username,
                    is_active=True,
                    is_superuser=True,
                    is_staff=True,
                ).first()

                if owner is not None and owner.check_password(password):
                    cache.delete(hidden_login_key)
                    clear_current_tenant()
                    login(request, owner, backend='django.contrib.auth.backends.ModelBackend')
                    request.session.pop('tenant_id', None)
                    request.session.pop(TENANT_DB_ALIAS_SESSION_KEY, None)
                    request.session[PLATFORM_OWNER_SESSION_KEY] = True
                    request.session[PLATFORM_OWNER_USERNAME_KEY] = owner.username
                    write_platform_audit(
                        event_type='platform_login',
                        actor_username=owner.username,
                        notes='دخول منصة الإدارة عبر صفحة تسجيل الدخول الموحدة',
                    )
                    messages.success(request, 'تم تسجيل دخول منصة الإدارة الفوقية بنجاح.')

                    if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()):
                        return redirect(next_url)
                    return redirect('platform_switch_tenant')

                cache.set(
                    hidden_login_key,
                    current_attempts + 1,
                    timeout=HIDDEN_PLATFORM_LOGIN_WINDOW_SECONDS,
                )
                _write_hidden_platform_login_failure_audit(
                    request,
                    username,
                    reason='invalid_hidden_platform_credentials',
                )
                form.add_error(None, 'بيانات الدخول غير صحيحة.')

            elif bool(tenant_id_raw) != bool(tenant_key):
                if not tenant_id_raw:
                    form.add_error('tenant_id', 'يرجى إدخال معرف المعرض أو ترك حقلي المعرض فارغين.')
                if not tenant_key:
                    form.add_error('tenant_key', 'يرجى إدخال كلمة مرور المعرض أو ترك حقلي المعرض فارغين.')

            else:
                tenant_id = normalize_tenant_id(tenant_id_raw)
                tenant_metadata = get_cached_tenant_metadata(tenant_id)
                if tenant_metadata is None or not tenant_metadata.get('is_active'):
                    form.add_error('tenant_id', 'معرف المعرض غير موجود أو غير نشط.')
                elif not is_valid_tenant_access_key(tenant_metadata, tenant_key):
                    form.add_error('tenant_key', 'كلمة مرور المعرض غير صحيحة.')
                else:
                    alias = ensure_tenant_connection(tenant_id)
                    set_current_tenant(tenant_id, alias)
                    user = User.objects.using(alias).filter(username=username).first()

                    if user is None:
                        form.add_error('username', 'اسم المستخدم غير موجود داخل هذا المعرض.')
                    elif not user.check_password(password):
                        form.add_error('password', 'كلمة مرور الحساب غير صحيحة.')
                    elif not user.is_active:
                        form.add_error(None, 'هذا الحساب غير نشط.')
                    else:
                        request.session.pop(PENDING_TENANT_LOGIN_SESSION_KEY, None)
                        request.session.pop(PENDING_TENANT_REGISTER_SESSION_KEY, None)
                        set_current_tenant(tenant_id, alias)
                        login(request, user, backend='sales.auth_backend.TenantModelBackend')
                        request.session['tenant_id'] = tenant_id
                        request.session[TENANT_DB_ALIAS_SESSION_KEY] = alias
                        write_platform_audit(
                            event_type='tenant_login',
                            tenant_id=tenant_id,
                            actor_username=user.username,
                            notes='تسجيل دخول بكلمة المرور',
                        )
                        if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()):
                            return redirect(next_url)
                        return redirect('home')
    else:
        form = TenantLoginForm()

    return render(
        request,
        'admin/login.html',
        {
            'form': form,
            'next': next_url,
            'google_oauth_enabled': _google_oauth_enabled(),
        },
    )


def google_auth_start(request):
    flow = (request.GET.get('flow') or '').strip().lower()
    if flow not in {'login', 'register'}:
        messages.error(request, 'نوع عملية Google غير معروف.')
        return redirect('login')

    if not _google_oauth_enabled():
        messages.error(request, 'Google OAuth غير مفعّل حاليًا في إعدادات النظام.')
        return redirect('login' if flow == 'login' else 'register')

    if flow == 'login':
        pending_login = request.session.get(PENDING_TENANT_LOGIN_SESSION_KEY)
        if not _is_recent_pending_payload(pending_login):
            request.session.pop(PENDING_TENANT_LOGIN_SESSION_KEY, None)
            messages.error(request, 'انتهت صلاحية خطوة كلمة المرور. أعد تسجيل الدخول.')
            return redirect('login')
    else:
        pending_register = request.session.get(PENDING_TENANT_REGISTER_SESSION_KEY)
        if not _is_recent_pending_payload(pending_register):
            request.session.pop(PENDING_TENANT_REGISTER_SESSION_KEY, None)
            messages.error(request, 'انتهت صلاحية خطوة إنشاء الحساب. أعد المحاولة.')
            return redirect('register')

    return redirect(_build_google_authorize_url(request, flow))


def google_auth_callback(request):
    oauth_state = request.session.get(GOOGLE_OAUTH_STATE_SESSION_KEY) or {}
    flow = oauth_state.get('flow')
    state = (request.GET.get('state') or '').strip()
    code = (request.GET.get('code') or '').strip()
    error = (request.GET.get('error') or '').strip()

    if error:
        _clear_pending_google_flow_session(request)
        messages.error(request, 'تم إلغاء المصادقة عبر Google أو فشلت العملية.')
        return redirect('login' if flow == 'login' else 'register')

    if not code or not state or state != oauth_state.get('token'):
        _clear_pending_google_flow_session(request)
        messages.error(request, 'طلب Google غير صالح. أعد المحاولة.')
        return redirect('login' if flow == 'login' else 'register')

    if flow not in {'login', 'register'}:
        _clear_pending_google_flow_session(request)
        messages.error(request, 'انتهت جلسة المصادقة. أعد المحاولة.')
        return redirect('login')

    try:
        token_payload = _exchange_google_code_for_tokens(code)
        id_token = (token_payload.get('id_token') or '').strip()
        if not id_token:
            raise ValidationError('تعذر استلام رمز تعريف Google.')
        google_profile = _verify_google_id_token(id_token, oauth_state.get('nonce'))
    except (ValidationError, urllib.error.URLError, TimeoutError, ValueError) as exc:
        _clear_pending_google_flow_session(request)
        messages.error(request, f'تعذر التحقق من Google: {exc}')
        return redirect('login' if flow == 'login' else 'register')

    if flow == 'login':
        pending_login = request.session.get(PENDING_TENANT_LOGIN_SESSION_KEY)
        if not _is_recent_pending_payload(pending_login):
            _clear_pending_google_flow_session(request)
            messages.error(request, 'انتهت صلاحية خطوة الدخول. أعد تسجيل الدخول.')
            return redirect('login')

        tenant_id = normalize_tenant_id(pending_login.get('tenant_id'))
        alias = (pending_login.get('tenant_alias') or '').strip()
        username = (pending_login.get('username') or '').strip()
        next_url = (pending_login.get('next_url') or '').strip()

        if not tenant_id or not alias.startswith('tenant_') or not username:
            _clear_pending_google_flow_session(request)
            messages.error(request, 'بيانات الجلسة ناقصة. أعد تسجيل الدخول.')
            return redirect('login')

        alias = ensure_tenant_connection(tenant_id)

        user = User.objects.using(alias).filter(username=username, is_active=True).first()
        if user is None:
            _clear_pending_google_flow_session(request)
            messages.error(request, 'الحساب غير موجود أو غير نشط.')
            return redirect('login')

        try:
            _bind_or_validate_google_identity(alias=alias, user=user, google_profile=google_profile)
        except ValidationError as exc:
            _clear_pending_google_flow_session(request)
            messages.error(request, str(exc))
            return redirect('login')

        set_current_tenant(tenant_id, alias)
        login(request, user, backend='sales.auth_backend.TenantModelBackend')
        request.session['tenant_id'] = tenant_id
        request.session[TENANT_DB_ALIAS_SESSION_KEY] = alias
        write_platform_audit(
            event_type='tenant_login',
            tenant_id=tenant_id,
            actor_username=user.username,
            notes='تسجيل دخول بكلمة المرور + Google',
        )
        _clear_pending_google_flow_session(request)
        if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()):
            return redirect(next_url)
        return redirect('home')

    pending_register = request.session.get(PENDING_TENANT_REGISTER_SESSION_KEY)
    if not _is_recent_pending_payload(pending_register):
        _clear_pending_google_flow_session(request)
        messages.error(request, 'انتهت صلاحية خطوة إنشاء الحساب. أعد المحاولة.')
        return redirect('register')

    tenant_id = normalize_tenant_id(pending_register.get('tenant_id'))
    tenant_key = pending_register.get('tenant_key') or ''
    showroom_name = pending_register.get('showroom_name') or ''
    username = pending_register.get('username') or ''
    password = pending_register.get('password') or ''

    try:
        with transaction.atomic(using='default'):
            tenant = PlatformTenant(name=showroom_name, tenant_id=tenant_id)
            tenant.set_access_key(tenant_key)
            tenant.save(using='default')
    except IntegrityError:
        _clear_pending_google_flow_session(request)
        messages.error(request, 'هذا المعرف مستخدم مسبقًا. اختر معرفًا آخر.')
        return redirect('register')

    try:
        alias = migrate_tenant_database(tenant_id)
        user = User.objects.db_manager(alias).create_superuser(
            username=username,
            email=google_profile['email'],
            password=password,
        )
        TenantUserGoogleIdentity.objects.using(alias).create(
            user=user,
            google_sub=google_profile['sub'],
            google_email=google_profile['email'],
            email_verified=True,
        )
    except Exception:
        tenant.delete(using='default')
        _clear_pending_google_flow_session(request)
        messages.error(request, 'تعذر إنشاء قاعدة بيانات المعرض أو ربط Google. حاول مرة أخرى.')
        return redirect('register')

    set_current_tenant(tenant_id, alias)
    login(request, user, backend='sales.auth_backend.TenantModelBackend')
    request.session['tenant_id'] = tenant_id
    request.session[TENANT_DB_ALIAS_SESSION_KEY] = alias
    _clear_pending_google_flow_session(request)
    messages.success(request, f'تم إنشاء بيئة معرض {showroom_name} وربط حساب Google بنجاح.')
    return redirect('home')


@ensure_csrf_cookie
def platform_owner_login(request):
    if request.method == 'POST':
        form = PlatformOwnerLoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            owner = User.objects.using('default').filter(username=username, is_active=True, is_superuser=True, is_staff=True).first()

            if owner is None or not owner.check_password(password):
                form.add_error(None, 'بيانات حساب المنصة غير صحيحة.')
            else:
                clear_current_tenant()
                login(request, owner, backend='django.contrib.auth.backends.ModelBackend')
                request.session.pop('tenant_id', None)
                request.session.pop(TENANT_DB_ALIAS_SESSION_KEY, None)
                request.session[PLATFORM_OWNER_SESSION_KEY] = True
                request.session[PLATFORM_OWNER_USERNAME_KEY] = owner.username
                write_platform_audit(
                    event_type='platform_login',
                    actor_username=owner.username,
                    notes='دخول منصة الإدارة الفوقية',
                )
                messages.success(request, 'تم تسجيل دخول منصة الإدارة الفوقية بنجاح.')
                return redirect('platform_switch_tenant')
    else:
        form = PlatformOwnerLoginForm()

    return render(request, 'admin/platform_login.html', {'form': form})


@login_required
def platform_switch_tenant(request):
    _require_platform_owner_session(request)
    owner_username = request.session.get(PLATFORM_OWNER_USERNAME_KEY)

    if request.method == 'POST':
        form = TenantSwitchForm(request.POST)
        if form.is_valid():
            tenant_id = normalize_tenant_id(form.cleaned_data['tenant_id'])
            tenant_metadata = get_cached_tenant_metadata(tenant_id)

            if tenant_metadata is None or not tenant_metadata.get('is_active'):
                form.add_error('tenant_id', 'المعرض غير موجود أو غير نشط.')
            else:
                alias = ensure_tenant_connection(tenant_id)
                set_current_tenant(tenant_id, alias)

                target_user = User.objects.using(alias).filter(is_superuser=True, is_active=True).order_by('id').first()
                if target_user is None:
                    target_user = User.objects.using(alias).filter(is_staff=True, is_active=True).order_by('id').first()
                if target_user is None:
                    target_user = User.objects.using(alias).filter(is_active=True).order_by('id').first()

                if target_user is None:
                    form.add_error('tenant_id', 'لا يوجد حسابات فعالة داخل هذا المعرض للدخول الفني.')
                else:
                    login(request, target_user, backend='sales.auth_backend.TenantModelBackend')
                    request.session['tenant_id'] = tenant_id
                    request.session[TENANT_DB_ALIAS_SESSION_KEY] = alias
                    request.session[PLATFORM_OWNER_SESSION_KEY] = True
                    if owner_username:
                        request.session[PLATFORM_OWNER_USERNAME_KEY] = owner_username
                    request.session['platform_owner_current_tenant'] = tenant_id
                    write_platform_audit(
                        event_type='platform_switch',
                        tenant_id=tenant_id,
                        actor_username=owner_username or target_user.username,
                        notes='تحويل فني إلى بيئة معرض',
                    )
                    messages.success(request, f"تم التحويل الفني إلى معرض: {tenant_metadata.get('name')}")
                    return redirect(f"{reverse('platform_switch_tenant')}?switched=1")
    else:
        form = TenantSwitchForm()

    tenants = PlatformTenant.objects.using('default').filter(is_active=True).order_by('name')
    return render(request, 'admin/platform_switch_tenant.html', {
        'form': form,
        'tenants': tenants,
        'switch_success_ok_url': reverse('home'),
        'exit_success_ok_url': reverse('platform_switch_tenant'),
    })


@login_required
@require_POST
def platform_exit_impersonation(request):
    _require_platform_owner_session(request)

    owner_username = request.session.get(PLATFORM_OWNER_USERNAME_KEY)
    owner = User.objects.using('default').filter(username=owner_username, is_active=True, is_superuser=True, is_staff=True).first()
    if owner is None:
        raise PermissionDenied

    clear_current_tenant()
    login(request, owner, backend='django.contrib.auth.backends.ModelBackend')
    request.session.pop('tenant_id', None)
    request.session.pop(TENANT_DB_ALIAS_SESSION_KEY, None)
    request.session[PLATFORM_OWNER_SESSION_KEY] = True
    request.session[PLATFORM_OWNER_USERNAME_KEY] = owner.username
    request.session.pop('platform_owner_current_tenant', None)
    write_platform_audit(
        event_type='platform_exit_switch',
        actor_username=owner.username,
        notes='إنهاء جلسة التحويل الفني',
    )
    messages.success(request, 'تم الرجوع إلى حساب منصة الإدارة الفوقية.')
    return redirect(f"{reverse('platform_switch_tenant')}?exited=1")


def user_logout(request):
    """Logs out the current user and send them to the login/registration screen."""
    tenant_id = normalize_tenant_id(request.session.get('tenant_id'))
    actor_username = request.user.username if getattr(request, 'user', None) and request.user.is_authenticated else ''

    if tenant_id:
        write_platform_audit(
            event_type='tenant_logout',
            tenant_id=tenant_id,
            actor_username=actor_username,
            notes='تسجيل خروج من بيئة المعرض',
        )

    logout(request)
    request.session.pop('tenant_id', None)
    request.session.pop(TENANT_DB_ALIAS_SESSION_KEY, None)
    request.session.pop(PLATFORM_OWNER_SESSION_KEY, None)
    request.session.pop(PLATFORM_OWNER_USERNAME_KEY, None)
    request.session.pop('platform_owner_current_tenant', None)
    clear_current_tenant()
    # بعد الخروج نعيد التوجيه إلى صفحة تسجيل الدخول/إنشاء حساب
    return redirect('login')


@login_required
@require_http_methods(['GET', 'POST'])
def user_theme_preference(request):
    tenant_id = normalize_tenant_id(request.session.get('tenant_id')) or ''
    username = (getattr(request.user, 'username', '') or '').strip()

    if not username:
        return JsonResponse({'error': 'invalid_user'}, status=400)

    if request.method == 'GET':
        preference = (
            UserThemePreference.objects.using('default')
            .filter(tenant_id=tenant_id, username=username)
            .first()
        )
        theme = preference.theme if preference else UserThemePreference.THEME_DARK
        updated_at = preference.updated_at.isoformat() if preference else None
        return JsonResponse({'theme': theme, 'updated_at': updated_at})

    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except (ValueError, UnicodeDecodeError):
        return JsonResponse({'error': 'invalid_payload'}, status=400)

    theme = (payload.get('theme') or '').strip().lower()
    if theme not in {
        UserThemePreference.THEME_DARK,
        UserThemePreference.THEME_LIGHT,
        UserThemePreference.THEME_AUTO,
    }:
        return JsonResponse({'error': 'invalid_theme'}, status=400)

    preference, _ = UserThemePreference.objects.using('default').update_or_create(
        tenant_id=tenant_id,
        username=username,
        defaults={'theme': theme},
    )
    return JsonResponse({'saved': True, 'theme': preference.theme, 'updated_at': preference.updated_at.isoformat()})


@ensure_csrf_cookie
def register(request):
    if request.method == 'POST':
        form = TenantRegisterForm(request.POST)
        if form.is_valid():
            tenant_id = normalize_tenant_id(form.cleaned_data['tenant_id'])
            tenant_key = form.cleaned_data['tenant_key']
            showroom_name = form.cleaned_data['showroom_name']
            username = form.cleaned_data['username']
            password = form.cleaned_data['password1']

            try:
                with transaction.atomic(using='default'):
                    tenant = PlatformTenant(name=showroom_name, tenant_id=tenant_id)
                    tenant.set_access_key(tenant_key)
                    tenant.save(using='default')
            except IntegrityError:
                form.add_error('tenant_id', 'هذا المعرف مستخدم مسبقًا. اختر معرفًا آخر.')
            else:
                try:
                    alias = migrate_tenant_database(tenant_id)
                    user = User.objects.db_manager(alias).create_superuser(
                        username=username,
                        email='',
                        password=password,
                    )
                except Exception:
                    tenant.delete(using='default')
                    form.add_error(None, 'تعذر إنشاء قاعدة بيانات المعرض أو حساب المدير. حاول مرة أخرى.')
                else:
                    request.session.pop(PENDING_TENANT_LOGIN_SESSION_KEY, None)
                    request.session.pop(PENDING_TENANT_REGISTER_SESSION_KEY, None)
                    set_current_tenant(tenant_id, alias)
                    login(request, user, backend='sales.auth_backend.TenantModelBackend')
                    request.session['tenant_id'] = tenant_id
                    request.session[TENANT_DB_ALIAS_SESSION_KEY] = alias
                    messages.success(request, f'تم إنشاء بيئة معرض {showroom_name} بنجاح.')
                    return redirect('home')
    else:
        form = TenantRegisterForm()

    return render(request, 'admin/register.html', {'form': form, 'google_oauth_enabled': _google_oauth_enabled()})


@require_interface_access('can_access_system_users')
def admin_user_filters(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    context = {
        'q': request.GET.get('q', ''),
        'is_staff': request.GET.get('is_staff__exact', ''),
        'is_superuser': request.GET.get('is_superuser__exact', ''),
        'is_active': request.GET.get('is_active__exact', ''),
    }
    return render(request, 'admin/auth/user/filters_search.html', context)


@require_interface_access('can_access_system_users')
def system_users(request):
    # عرض جميع مستخدمي النظام في صفحة منفصلة بصيغة جدول
    if not request.user.is_staff:
        raise PermissionDenied

    users = User.objects.all().order_by('-date_joined')
    return render(request, 'sales/system_users.html', {'users': users})


@require_interface_access('can_access_cars')
def available_cars_table(request):
    if not request.user.is_staff:
        raise PermissionDenied

    cars = Car.objects.filter(is_sold=False).order_by('-created_at')
    return render(request, 'sales/available_cars.html', {'cars': cars})


@require_interface_access('can_access_cars')
def sold_cars_cards(request):
    if not request.user.is_staff:
        raise PermissionDenied

    cars = Car.objects.filter(is_sold=True).order_by('-created_at')
    return render(request, 'sales/sold_cars_cards.html', {'cars': cars})


@require_interface_access('can_access_cars')
def sold_car_details(request, car_id):
    if not request.user.is_staff:
        raise PermissionDenied

    car = get_object_or_404(Car, id=car_id, is_sold=True)
    sale = getattr(car, 'sale', None)
    maintenance_records = car.maintenance_records.select_related('added_by').all()
    real_profit = None
    if sale is not None:
        real_profit = sale.sale_price - car.total_cost_price

    return render(request, 'sales/sold_car_details.html', {
        'car': car,
        'sale': sale,
        'maintenance_records': maintenance_records,
        'maintenance_total': car.maintenance_total,
        'total_cost_price': car.total_cost_price,
        'real_profit': real_profit,
    })


@require_interface_access('can_access_system_users')
def permissions_management(request):
    if not request.user.is_superuser:
        raise PermissionDenied

    users = User.objects.all().order_by('username')
    superusers_count = User.objects.filter(is_superuser=True).count()
    for account in users:
        InterfaceAccess.objects.get_or_create(user=account)

    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        target_user = get_object_or_404(User, id=user_id)
        access, _ = InterfaceAccess.objects.get_or_create(user=target_user)

        for field_name in PERMISSION_FIELDS:
            setattr(access, field_name, field_name in request.POST)

        access.save(update_fields=PERMISSION_FIELDS)
        return redirect(f"{reverse('permissions_management')}?saved=1")

    users_with_access = []
    for account in users:
        access, _ = InterfaceAccess.objects.get_or_create(user=account)
        is_current_user = account.id == request.user.id
        is_last_superuser = account.is_superuser and superusers_count <= 1
        can_delete = not is_current_user and not is_last_superuser
        users_with_access.append({
            'account': account,
            'access': access,
            'can_delete': can_delete,
            'is_current_user': is_current_user,
            'is_last_superuser': is_last_superuser,
        })

    context = {
        'users_with_access': users_with_access,
        'permissions_ok_url': reverse('permissions_management'),
    }
    return render(request, 'sales/permissions_management.html', context)


@require_interface_access('can_access_system_users')
@require_POST
def delete_system_user(request, user_id):
    if not request.user.is_superuser:
        raise PermissionDenied

    target_user = get_object_or_404(User, id=user_id)
    if target_user.id == request.user.id:
        return redirect('permissions_management')

    if target_user.is_superuser and User.objects.filter(is_superuser=True).count() <= 1:
        return redirect('permissions_management')

    target_user.delete()
    return redirect(f"{reverse('permissions_management')}?deleted=1")